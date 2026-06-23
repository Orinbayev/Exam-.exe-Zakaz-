"""
Fan (Subject) boshqaruvi — fanlar va ularning savollarini bir joyda.
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QListWidget, QListWidgetItem, QTableWidget, QTableWidgetItem,
    QDialog, QLineEdit, QComboBox, QTextEdit, QInputDialog,
    QMessageBox, QHeaderView, QFrame, QSizePolicy,
    QScrollArea, QFileDialog, QProgressDialog,
)
from PyQt6.QtCore import Qt, QSize, QTimer
from ...worker import ApiWorker
from PyQt6.QtGui import QFont, QColor
from ...api_client import api, APIError
from ...styles import COLORS
from ...i18n import t


# ─────────────────────────────────────────────────────────────────────────────
# Stil konstantalari
# ─────────────────────────────────────────────────────────────────────────────

_DLG_BASE = f"""
QDialog {{
    background: {COLORS['bg_medium']};
    color: white;
    font-family: 'Segoe UI', Arial;
}}
QLabel {{ color: white; background: transparent; }}
QLineEdit, QTextEdit, QComboBox {{
    background: {COLORS['bg_dark']};
    color: white;
    border: 1.5px solid {COLORS['border']};
    border-radius: 8px;
    padding: 8px 12px;
    font-size: 13px;
}}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus {{
    border: 2px solid {COLORS['primary_light']};
}}
QComboBox::drop-down {{ border: none; width: 24px; }}
QPushButton {{
    background: {COLORS['primary']};
    color: white; border: none; border-radius: 8px;
    padding: 9px 20px; font-size: 13px; font-weight: bold;
    min-height: 36px;
}}
QPushButton:hover {{ background: {COLORS['primary_light']}; }}
QPushButton#cancel {{
    background: {COLORS['bg_light']};
    border: 1px solid {COLORS['border']};
}}
QPushButton#cancel:hover {{ background: {COLORS['bg_dark']}; }}
"""

_PANEL_STYLE = f"""
    QFrame {{
        background: {COLORS['bg_medium']};
        border: 1px solid {COLORS['border']};
        border-radius: 12px;
    }}
"""

_LIST_STYLE = f"""
    QListWidget {{
        background: {COLORS['bg_dark']};
        border: 1px solid {COLORS['border']};
        border-radius: 8px;
        color: white;
        font-size: 13px;
        outline: none;
    }}
    QListWidget::item {{
        padding: 11px 14px;
        border-bottom: 1px solid {COLORS['border']};
    }}
    QListWidget::item:selected {{
        background: {COLORS['primary']};
        color: white;
        border-radius: 6px;
    }}
    QListWidget::item:hover:!selected {{
        background: {COLORS['bg_light']};
    }}
"""

_TABLE_STYLE = f"""
    QTableWidget {{
        background: {COLORS['bg_dark']};
        color: white;
        border: 1px solid {COLORS['border']};
        border-radius: 8px;
        gridline-color: {COLORS['border']};
        font-size: 13px;
        outline: none;
    }}
    QTableWidget::item {{ padding: 6px 10px; }}
    QTableWidget::item:selected {{
        background: {COLORS['primary']};
        color: white;
    }}
    QHeaderView::section {{
        background: {COLORS['bg_light']};
        color: {COLORS['text_secondary']};
        padding: 9px 10px;
        border: none;
        border-right: 1px solid {COLORS['border']};
        border-bottom: 1px solid {COLORS['border']};
        font-weight: bold;
        font-size: 12px;
    }}
"""


# ─────────────────────────────────────────────────────────────────────────────
# Fan qo'shish dialogi
# ─────────────────────────────────────────────────────────────────────────────

class FanDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(t("fan.dialog_title"))
        self.setFixedSize(440, 290)
        self.setStyleSheet(f"""
            QDialog {{
                background: {COLORS['bg_medium']};
                border-radius: 16px;
            }}
        """)
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Header ─────────────────────────────────────────────────────────────
        header = QFrame()
        header.setFixedHeight(100)
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {COLORS['primary_dark']}, stop:1 {COLORS['primary']});
                border-radius: 0px;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
            }}
        """)
        h_lay = QVBoxLayout(header)
        h_lay.setContentsMargins(0, 14, 0, 14)
        h_lay.setSpacing(4)

        ico = QLabel("📚")
        ico.setFont(QFont("Segoe UI Emoji", 26))
        ico.setAlignment(Qt.AlignmentFlag.AlignCenter)
        h_lay.addWidget(ico)

        title_lbl = QLabel(t("fan.dialog_title"))
        title_lbl.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_lbl.setStyleSheet("color: white; background: transparent;")
        h_lay.addWidget(title_lbl)
        root.addWidget(header)

        # ── Body ───────────────────────────────────────────────────────────────
        body = QFrame()
        body.setStyleSheet(f"""
            QFrame {{
                background: {COLORS['bg_medium']};
                border-bottom-left-radius: 10px;
                border-bottom-right-radius: 10px;
            }}
        """)
        b_lay = QVBoxLayout(body)
        b_lay.setContentsMargins(28, 20, 28, 24)
        b_lay.setSpacing(14)

        lbl = QLabel(t("fan.dialog_label"))
        lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
        lbl.setStyleSheet(f"color: {COLORS['text_secondary']}; background: transparent;")
        b_lay.addWidget(lbl)

        self.inp = QLineEdit()
        self.inp.setPlaceholderText(t("fan.dialog_ph"))
        self.inp.setFixedHeight(46)
        self.inp.setFont(QFont("Segoe UI", 12))
        self.inp.setStyleSheet(f"""
            QLineEdit {{
                background: {COLORS['bg_dark']};
                color: white;
                border: 2px solid {COLORS['border']};
                border-radius: 10px;
                padding: 0 14px;
                font-size: 13px;
            }}
            QLineEdit:focus {{
                border: 2px solid {COLORS['primary_light']};
                background: #0D2137;
            }}
        """)
        self.inp.returnPressed.connect(self._ok)
        b_lay.addWidget(self.inp)

        b_lay.addSpacing(4)

        # ── Buttons ────────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)

        cl_btn = QPushButton(t("fan.dialog_cancel"))
        cl_btn.setFixedSize(140, 42)
        cl_btn.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        cl_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                color: {COLORS['text_secondary']};
                border: 2px solid {COLORS['border']};
                border-radius: 10px;
                font-size: 12px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: {COLORS['bg_light']};
                color: white;
                border-color: {COLORS['primary_light']};
            }}
        """)
        cl_btn.clicked.connect(self.reject)

        ok_btn = QPushButton(t("fan.dialog_ok"))
        ok_btn.setFixedSize(160, 42)
        ok_btn.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        ok_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['success']}, stop:1 #388E3C);
                color: white;
                border: none;
                border-radius: 10px;
                font-size: 12px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #388E3C, stop:1 {COLORS['success_light']});
            }}
            QPushButton:pressed {{
                background: {COLORS['success']};
            }}
        """)
        ok_btn.clicked.connect(self._ok)

        btn_row.addStretch()
        btn_row.addWidget(cl_btn)
        btn_row.addWidget(ok_btn)
        btn_row.addStretch()
        b_lay.addLayout(btn_row)

        root.addWidget(body)

    def _ok(self):
        if not self.inp.text().strip():
            self.inp.setPlaceholderText(t("fan.empty_warn"))
            self.inp.setStyleSheet(self.inp.styleSheet() +
                "border: 2px solid #EF5350 !important;")
            return
        self.accept()

    def get_name(self):
        return self.inp.text().strip()


# ─────────────────────────────────────────────────────────────────────────────
# Savol qo'shish / tahrirlash dialogi
# ─────────────────────────────────────────────────────────────────────────────

class QuestionDialog(QDialog):
    def __init__(self, fan_name: str = "", question: dict = None, parent=None):
        super().__init__(parent)
        self._q = question
        _action = t("q.dialog_add") if not question else t("q.dialog_edit")
        self.setWindowTitle(_action + (f" — {fan_name}" if fan_name else ""))
        self.setFixedWidth(540)
        self.setStyleSheet(_DLG_BASE)
        self._build()
        if question:
            self._fill(question)

    def _build(self):
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")

        inner = QWidget()
        inner.setStyleSheet("background: transparent;")
        lay = QVBoxLayout(inner)
        lay.setContentsMargins(26, 20, 26, 20)
        lay.setSpacing(12)

        # Header icon
        ico = QLabel("❓" if not self._q else "✏️")
        ico.setFont(QFont("Segoe UI Emoji", 24))
        ico.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(ico)

        def _lbl(text):
            l = QLabel(text)
            l.setFont(QFont("Segoe UI", 10, QFont.Weight.DemiBold))
            l.setStyleSheet(f"color: {COLORS['text_secondary']};")
            return l

        # Savol matni
        lay.addWidget(_lbl(t("q.text_label")))
        self.txt = QTextEdit()
        self.txt.setFixedHeight(90)
        self.txt.setPlaceholderText(t("q.text_ph"))
        lay.addWidget(self.txt)

        # Variantlar
        self.opts: list[QLineEdit] = []
        labels = ["A  variant *", "B  variant *", "C  variant *", "D  variant *"]
        colors = ["#66BB6A", "#42A5F5", "#FFA726", "#EF5350"]
        for lbl_txt, clr in zip(labels, colors):
            row = QHBoxLayout()
            row.setSpacing(8)
            badge = QLabel(lbl_txt[0])
            badge.setFixedSize(28, 28)
            badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
            badge.setStyleSheet(f"""
                background: {clr}; color: white;
                border-radius: 6px; font-weight: bold; font-size: 13px;
            """)
            inp = QLineEdit()
            inp.setPlaceholderText(lbl_txt)
            inp.setFixedHeight(40)
            row.addWidget(badge)
            row.addWidget(inp, stretch=1)
            lay.addLayout(row)
            self.opts.append(inp)

        # To'g'ri javob
        lay.addWidget(_lbl(t("q.correct_label")))
        self.answer = QComboBox()
        self.answer.addItems([
            "A  — birinchi variant",
            "B  — ikkinchi variant",
            "C  — uchinchi variant",
            "D  — to'rtinchi variant",
        ])
        self.answer.setFixedHeight(48)
        self.answer.setStyleSheet(f"""
            QComboBox {{
                background: {COLORS['bg_dark']};
                color: white;
                border: 2px solid {COLORS['primary']};
                border-radius: 10px;
                padding: 0 14px;
                font-size: 14px;
                font-weight: bold;
                font-family: 'Segoe UI', Arial;
            }}
            QComboBox::drop-down {{
                border: none;
                width: 32px;
            }}
            QComboBox::down-arrow {{
                border-left:  6px solid transparent;
                border-right: 6px solid transparent;
                border-top:   7px solid rgba(255,255,255,0.85);
                margin-right: 10px;
            }}
            QComboBox QAbstractItemView {{
                background: {COLORS['bg_medium']};
                color: white;
                selection-background-color: {COLORS['primary']};
                border: 1px solid {COLORS['border']};
                font-size: 13px;
                padding: 4px;
                outline: none;
            }}
        """)
        lay.addWidget(self.answer)

        # Tugmalar
        lay.addSpacing(6)
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_row.addStretch()
        cancel = QPushButton(t("fan.dialog_cancel"))
        cancel.setObjectName("cancel")
        cancel.setFixedSize(110, 34)
        cancel.clicked.connect(self.reject)
        save = QPushButton(t("fan.dialog_ok"))
        save.setFixedSize(110, 34)
        save.clicked.connect(self._ok)
        btn_row.addWidget(cancel)
        btn_row.addWidget(save)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        scroll.setWidget(inner)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

    def _fill(self, q: dict):
        self.txt.setPlainText(q.get("text", ""))
        self.opts[0].setText(q.get("option_a", ""))
        self.opts[1].setText(q.get("option_b", ""))
        self.opts[2].setText(q.get("option_c", ""))
        self.opts[3].setText(q.get("option_d", ""))
        ans = q.get("correct_answer", "A").upper()
        self.answer.setCurrentIndex({"A": 0, "B": 1, "C": 2, "D": 3}.get(ans, 0))

    def _ok(self):
        if not self.txt.toPlainText().strip():
            QMessageBox.warning(self, t("common.warning"), t("q.empty_warn"))
            return
        for i, opt in enumerate(self.opts):
            if not opt.text().strip():
                QMessageBox.warning(self, t("common.warning"),
                    f"{chr(65+i)} {t('q.opt_warn')}")
                return
        self.accept()

    def get_data(self, category_id: int = None) -> dict:
        # "A  — birinchi variant" dan faqat "A" ni olamiz
        ans_letter = self.answer.currentText().strip()[0].upper()
        return {
            "text":           self.txt.toPlainText().strip(),
            "option_a":       self.opts[0].text().strip(),
            "option_b":       self.opts[1].text().strip(),
            "option_c":       self.opts[2].text().strip(),
            "option_d":       self.opts[3].text().strip(),
            "correct_answer": ans_letter,
            "category_id":    category_id,
            "difficulty":     "medium",
        }


# ─────────────────────────────────────────────────────────────────────────────
# Asosiy widget
# ─────────────────────────────────────────────────────────────────────────────

class FanWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._fans: list = []
        self._questions: list = []
        self._fan_id: int | None = None
        self._fan_name: str = ""
        self._setup_ui()
        self.refresh()

    # ── UI qurilishi ──────────────────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(14)

        # ── Sarlavha ──────────────────────────────────────────────────────────
        hdr = QHBoxLayout()
        title = QLabel(t("fan.title"))
        title.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
        hdr.addWidget(title)
        hdr.addStretch()

        hint_lbl = QLabel(t("fan.hint"))
        hint_lbl.setFont(QFont("Segoe UI", 11))
        hint_lbl.setStyleSheet(f"color: {COLORS['text_secondary']};")
        hdr.addWidget(hint_lbl)
        root.addLayout(hdr)

        # ── Asosiy ikki panel ─────────────────────────────────────────────────
        panels = QHBoxLayout()
        panels.setSpacing(14)

        # ── CHAP: Fan ro'yxati ────────────────────────────────────────────────
        left = QFrame()
        left.setStyleSheet(_PANEL_STYLE)
        left.setFixedWidth(250)
        ll = QVBoxLayout(left)
        ll.setContentsMargins(12, 14, 12, 14)
        ll.setSpacing(10)

        l_hdr = QHBoxLayout()
        l_title = QLabel(t("fan.section_fans"))
        l_title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        l_hdr.addWidget(l_title)
        self._fan_count_lbl = QLabel(t("fan.count", n=0))
        self._fan_count_lbl.setStyleSheet(
            f"color: {COLORS['text_secondary']}; font-size: 11px;"
        )
        l_hdr.addStretch()
        l_hdr.addWidget(self._fan_count_lbl)
        ll.addLayout(l_hdr)

        self._fan_list = QListWidget()
        self._fan_list.setStyleSheet(_LIST_STYLE)
        self._fan_list.currentItemChanged.connect(self._on_fan_selected)
        ll.addWidget(self._fan_list)

        add_fan_btn = QPushButton(t("fan.add_btn"))
        add_fan_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['success']};
                color: white; border: none; border-radius: 8px;
                padding: 10px; font-size: 13px; font-weight: bold;
            }}
            QPushButton:hover {{ background: {COLORS['success_light']}; }}
        """)
        add_fan_btn.clicked.connect(self._add_fan)
        ll.addWidget(add_fan_btn)

        time_btn = QPushButton(t("fan.time_btn"))
        time_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['accent']};
                color: white; border: none; border-radius: 8px;
                padding: 8px; font-size: 12px; font-weight: bold;
            }}
            QPushButton:hover {{ background: {COLORS['accent_light']}; }}
        """)
        time_btn.clicked.connect(self._set_time_limit)
        ll.addWidget(time_btn)

        del_fan_btn = QPushButton(t("fan.del_btn"))
        del_fan_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['bg_dark']};
                color: {COLORS['danger_light']};
                border: 1px solid {COLORS['danger']};
                border-radius: 8px; padding: 8px; font-size: 12px;
            }}
            QPushButton:hover {{
                background: {COLORS['danger']}; color: white;
            }}
        """)
        del_fan_btn.clicked.connect(self._delete_fan)
        ll.addWidget(del_fan_btn)

        panels.addWidget(left)

        # ── O'NG: Savollar ────────────────────────────────────────────────────
        right = QFrame()
        right.setStyleSheet(_PANEL_STYLE)
        rl = QVBoxLayout(right)
        rl.setContentsMargins(14, 14, 14, 14)
        rl.setSpacing(10)

        # O'ng sarlavha
        r_hdr = QHBoxLayout()
        self._q_title = QLabel(t("fan.q_section"))
        self._q_title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        r_hdr.addWidget(self._q_title)
        r_hdr.addStretch()
        self._q_count_lbl = QLabel("")
        self._q_count_lbl.setStyleSheet(
            f"color: {COLORS['text_secondary']}; font-size: 11px;"
        )
        r_hdr.addWidget(self._q_count_lbl)
        rl.addLayout(r_hdr)

        # Statistika qatori
        self._stat_row = QHBoxLayout()
        self._stat_active_lbl = QLabel("")
        self._stat_frozen_lbl = QLabel("")
        for lbl in (self._stat_active_lbl, self._stat_frozen_lbl):
            lbl.setFont(QFont("Segoe UI", 11))
            lbl.setStyleSheet("background: transparent;")
        self._stat_row.addWidget(self._stat_active_lbl)
        self._stat_row.addSpacing(16)
        self._stat_row.addWidget(self._stat_frozen_lbl)
        self._stat_row.addStretch()
        rl.addLayout(self._stat_row)

        # Bo'sh holat
        self._empty_lbl = QLabel(t("fan.empty_select"))
        self._empty_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_lbl.setStyleSheet(
            f"color: {COLORS['text_secondary']}; font-size: 14px; background: transparent;"
        )

        # Jadval
        self._tbl = QTableWidget()
        self._tbl.setColumnCount(6)
        self._tbl.setHorizontalHeaderLabels(
            [t("fan.tbl_num"), t("fan.tbl_text"), "A", "B", "C", "D"]
        )
        self._tbl.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self._tbl.setColumnWidth(0, 40)
        self._tbl.setColumnWidth(2, 130)
        self._tbl.setColumnWidth(3, 130)
        self._tbl.setColumnWidth(4, 130)
        self._tbl.setColumnWidth(5, 130)
        self._tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._tbl.setAlternatingRowColors(True)
        self._tbl.setStyleSheet(_TABLE_STYLE + """
            QTableWidget { alternate-background-color: rgba(255,255,255,0.03); }
        """)
        self._tbl.hide()

        rl.addWidget(self._empty_lbl, stretch=1)
        rl.addWidget(self._tbl, stretch=1)

        # Tugmalar qatori
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        _disabled_style = f"""
            QPushButton:disabled {{
                background: {COLORS['bg_light']};
                color: {COLORS['text_secondary']};
            }}
        """

        self._add_q_btn = QPushButton(t("fan.add_q_btn"))
        self._add_q_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['success']};
                color: white; border: none; border-radius: 8px;
                padding: 9px 16px; font-size: 13px; font-weight: bold;
            }}
            QPushButton:hover {{ background: {COLORS['success_light']}; }}
        """ + _disabled_style)
        self._add_q_btn.setEnabled(False)
        self._add_q_btn.clicked.connect(self._add_question)

        self._excel_import_btn = QPushButton(t("fan.excel_btn"))
        self._excel_import_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['accent']};
                color: white; border: none; border-radius: 8px;
                padding: 9px 14px; font-size: 12px; font-weight: bold;
            }}
            QPushButton:hover {{ background: {COLORS['accent_light']}; }}
        """ + _disabled_style)
        self._excel_import_btn.setEnabled(False)
        self._excel_import_btn.setToolTip(t("fan.excel_btn"))
        self._excel_import_btn.clicked.connect(self._import_excel)

        self._excel_tpl_btn = QPushButton(t("fan.excel_tpl"))
        self._excel_tpl_btn.setObjectName("secondary")
        self._excel_tpl_btn.setToolTip(t("fan.excel_tpl"))
        self._excel_tpl_btn.clicked.connect(self._download_template)

        self._show_frozen_btn = QPushButton(t("fan.frozen_btn"))
        self._show_frozen_btn.setObjectName("secondary")
        self._show_frozen_btn.setCheckable(True)
        self._show_frozen_btn.clicked.connect(self._toggle_show_frozen)

        btn_row.addWidget(self._add_q_btn)
        btn_row.addWidget(self._excel_import_btn)
        btn_row.addWidget(self._excel_tpl_btn)
        btn_row.addStretch()
        btn_row.addWidget(self._show_frozen_btn)
        rl.addLayout(btn_row)

        panels.addWidget(right, stretch=1)
        root.addLayout(panels, stretch=1)

        self._show_all_frozen = False

    # ── Fanlar ────────────────────────────────────────────────────────────────

    def refresh(self):
        self._fan_count_lbl.setText(t("fan.loading"))
        self._w_fans = ApiWorker(api.get_categories)
        self._w_fans.done.connect(self._apply_fans)
        self._w_fans.error.connect(lambda _: self._apply_fans([]))
        self._w_fans.start()

    def _apply_fans(self, fans):
        self._fans = fans
        self._fan_list.blockSignals(True)
        self._fan_list.clear()
        for f in self._fans:
            q_count = f.get("question_count", 0)
            t_lim = f.get("time_limit", 30)
            badge = t("fan.badge", q=q_count, t=t_lim)
            item = QListWidgetItem(f"📚  {f['name']}{badge}")
            item.setSizeHint(QSize(0, 48))
            item.setData(Qt.ItemDataRole.UserRole, f["id"])
            item.setData(Qt.ItemDataRole.UserRole + 1, f["name"])
            item.setData(Qt.ItemDataRole.UserRole + 2, t_lim)
            self._fan_list.addItem(item)
        self._fan_list.blockSignals(False)
        self._fan_count_lbl.setText(t("fan.count", n=len(self._fans)))

        if self._fan_id:
            for i in range(self._fan_list.count()):
                if self._fan_list.item(i).data(Qt.ItemDataRole.UserRole) == self._fan_id:
                    self._fan_list.setCurrentRow(i)
                    break

    def _on_fan_selected(self, item):
        if not item:
            return
        self._fan_id = item.data(Qt.ItemDataRole.UserRole)
        self._fan_name = item.data(Qt.ItemDataRole.UserRole + 1)
        self._q_title.setText(t("fan.q_title", name=self._fan_name))
        self._add_q_btn.setEnabled(True)
        self._excel_import_btn.setEnabled(True)
        self._load_questions()

    def _add_fan(self):
        dlg = FanDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            try:
                api.create_category(dlg.get_name())
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, t("common.error"), str(e))

    def _set_time_limit(self):
        item = self._fan_list.currentItem()
        if not item:
            QMessageBox.information(self, t("fan.select_warn"), t("fan.select_hint"))
            return
        fan = next((f for f in self._fans if f["id"] == self._fan_id), None)
        current_time = fan.get("time_limit", 30) if fan else 30

        from PyQt6.QtWidgets import QInputDialog
        val, ok = QInputDialog.getInt(
            self,
            t("fan.time_btn"),
            t("fan.time_input", name=self._fan_name, cur=current_time),
            current_time, 1, 300, 1
        )
        if not ok:
            return
        try:
            api.set_category_time_limit(self._fan_id, val)
            QMessageBox.information(
                self, t("fan.save_title"),
                t("fan.time_saved", name=self._fan_name, val=val)
            )
            self.refresh()
        except APIError as e:
            QMessageBox.critical(self, t("common.error"), str(e))

    def _delete_fan(self):
        item = self._fan_list.currentItem()
        if not item:
            QMessageBox.information(self, t("fan.select_warn"), t("fan.select_hint"))
            return
        name = item.data(Qt.ItemDataRole.UserRole + 1)
        r = QMessageBox.question(
            self, t("fan.del_title"),
            t("fan.del_confirm", name=name),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if r == QMessageBox.StandardButton.Yes:
            try:
                api.delete_category(self._fan_id)
                self._fan_id = None
                self._fan_name = ""
                self._q_title.setText(t("fan.q_section"))
                self._q_count_lbl.setText("")
                self._tbl.hide()
                self._empty_lbl.setText(t("fan.empty_select"))
                self._empty_lbl.show()
                self._add_q_btn.setEnabled(False)
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, t("common.error"), str(e))

    # ── Savollar ──────────────────────────────────────────────────────────────

    def _load_questions(self):
        if not self._fan_id:
            return
        try:
            self._questions = api.get_questions(category_id=self._fan_id)
        except APIError as e:
            QMessageBox.warning(self, t("common.error"), str(e))
            return
        self._render_table()

    def _toggle_show_frozen(self):
        self._show_all_frozen = self._show_frozen_btn.isChecked()
        btn_text = (t("fan.frozen_show_all") if self._show_all_frozen
                    else t("fan.frozen_btn"))
        self._show_frozen_btn.setText(btn_text)
        self._render_table()

    def _render_table(self):
        # Filterlash: agar "show frozen" yoqilmagan bo'lsa — faqat aktiv
        if self._show_all_frozen:
            rows = self._questions
        else:
            rows = [q for q in self._questions if q.get("is_active", True)]

        active_n = sum(1 for q in self._questions if q.get("is_active", True))
        frozen_n = len(self._questions) - active_n

        self._q_count_lbl.setText(t("fan.q_count_txt", n=len(self._questions)))
        self._stat_active_lbl.setText(t("fan.stat_active", n=active_n))
        self._stat_active_lbl.setStyleSheet(
            f"color: {COLORS['success_light']}; background: transparent;"
        )
        self._stat_frozen_lbl.setText(t("fan.stat_frozen", n=frozen_n))
        self._stat_frozen_lbl.setStyleSheet(
            f"color: {COLORS['text_secondary']}; background: transparent;"
        )

        if not rows:
            self._tbl.hide()
            if not self._questions:
                self._empty_lbl.setText(t("fan.empty_no_q", name=self._fan_name))
            else:
                self._empty_lbl.setText(t("fan.empty_frozen"))
            self._empty_lbl.show()
            return

        self._empty_lbl.hide()
        self._tbl.show()
        self._tbl.setRowCount(len(rows))

        for row_idx, q in enumerate(rows):
            self._tbl.setRowHeight(row_idx, 46)
            is_active = q.get("is_active", True)

            # # raqam
            num_item = QTableWidgetItem(str(row_idx + 1))
            num_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if not is_active:
                num_item.setForeground(QColor(COLORS["text_secondary"]))
            self._tbl.setItem(row_idx, 0, num_item)

            # Savol matni (qisqartirilgan)
            text = q.get("text", "")
            short = text[:80] + "..." if len(text) > 80 else text
            if not is_active:
                short = "❄️  " + short
            txt_item = QTableWidgetItem(short)
            txt_item.setFont(QFont("Segoe UI", 12))
            if not is_active:
                txt_item.setForeground(QColor(COLORS["text_secondary"]))
            txt_item.setToolTip(text)
            self._tbl.setItem(row_idx, 1, txt_item)

            # A B C D variantlari
            ans = q.get("correct_answer", "")
            opt_keys = ["option_a", "option_b", "option_c", "option_d"]
            opt_labels = ["A", "B", "C", "D"]
            opt_color = "#66BB6A"  # to'g'ri javob rangi

            for ci, (key, lbl) in enumerate(zip(opt_keys, opt_labels)):
                val = q.get(key, "")
                short_val = val[:22] + "…" if len(val) > 22 else val
                is_correct = (ans == lbl)
                item = QTableWidgetItem(
                    f"✔ {short_val}" if is_correct else short_val
                )
                item.setFont(QFont("Segoe UI", 11))
                item.setToolTip(val)
                if not is_active:
                    item.setForeground(QColor(COLORS["text_secondary"]))
                elif is_correct:
                    item.setForeground(QColor(opt_color))
                    item.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
                self._tbl.setItem(row_idx, 2 + ci, item)

        # Amallar ustunini alohida qo'shamiz (7-ustun)
        # Avval ustun sonini 7 ga o'rnatamiz
        self._tbl.setColumnCount(7)
        self._tbl.setHorizontalHeaderLabels(
            [t("fan.tbl_num"), t("fan.tbl_text"), "A", "B", "C", "D", t("fan.tbl_actions")]
        )
        self._tbl.setColumnWidth(6, 120)
        self._tbl.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )

        for row_idx, q in enumerate(rows):
            is_active = q.get("is_active", True)
            cell_w = QWidget()
            cell_w.setStyleSheet("background: transparent;")
            cly = QHBoxLayout(cell_w)
            cly.setContentsMargins(4, 3, 4, 3)
            cly.setSpacing(4)
            cly.setAlignment(Qt.AlignmentFlag.AlignCenter)

            ed = QPushButton("✏️")
            ed.setFixedSize(34, 30)
            ed.setObjectName("table_action")
            ed.setToolTip(t("fan.tip_edit"))
            ed.clicked.connect(lambda _, qid=q["id"]: self._edit_question(qid))

            frz = QPushButton("❄️" if is_active else "✅")
            frz.setFixedSize(34, 30)
            frz.setObjectName("table_action")
            frz.setToolTip(t("fan.tip_freeze") if is_active else t("fan.tip_activate"))
            frz.clicked.connect(lambda _, qid=q["id"]: self._toggle_question(qid))

            dl = QPushButton("🗑️")
            dl.setFixedSize(34, 30)
            dl.setObjectName("table_action_danger")
            dl.setToolTip(t("fan.tip_delete"))
            dl.clicked.connect(lambda _, qid=q["id"]: self._delete_question(qid))

            cly.addWidget(ed)
            cly.addWidget(frz)
            cly.addWidget(dl)
            self._tbl.setCellWidget(row_idx, 6, cell_w)

    # ── Savol CRUD ────────────────────────────────────────────────────────────

    def _add_question(self):
        dlg = QuestionDialog(fan_name=self._fan_name, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            try:
                api.create_question(dlg.get_data(category_id=self._fan_id))
                self._load_questions()
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, t("common.error"), str(e))

    def _edit_question(self, qid: int):
        q = next((x for x in self._questions if x["id"] == qid), None)
        if not q:
            return
        dlg = QuestionDialog(fan_name=self._fan_name, question=q, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            try:
                data = dlg.get_data(category_id=self._fan_id)
                data.pop("category_id", None)  # PUT uchun kerak emas
                api.update_question(qid, data)
                self._load_questions()
            except APIError as e:
                QMessageBox.critical(self, t("common.error"), str(e))

    def _toggle_question(self, qid: int):
        try:
            result = api.toggle_question_active(qid)
            q = next((x for x in self._questions if x["id"] == qid), None)
            if q:
                q["is_active"] = result.get("is_active", not q.get("is_active", True))
            self._render_table()
        except APIError as e:
            QMessageBox.critical(self, t("common.error"), str(e))

    def _delete_question(self, qid: int):
        q = next((x for x in self._questions if x["id"] == qid), None)
        if not q:
            return
        short = q["text"][:60] + "..." if len(q["text"]) > 60 else q["text"]
        r = QMessageBox.question(
            self, t("q.del_title"),
            t("fan.del_q_confirm", text=short),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if r == QMessageBox.StandardButton.Yes:
            try:
                api.delete_question(qid)
                self._load_questions()
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, t("common.error"), str(e))

    # ── Excel import / namuna ────────────────────────────────────────────────

    def _download_template(self):
        """Namuna Excel faylini yuklash."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(self, t("common.error"), "openpyxl kutubxonasi topilmadi!")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, t("fan.tpl_save_title"),
            t("fan.tpl_save_ph"), "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Savollar"

        headers = [
            "Savol matni",
            "A variant",
            "B variant",
            "C variant",
            "D variant",
            "To'g'ri javob (A/B/C/D)",
            "Daraja (easy / medium / hard)",
        ]
        hdr_fill = PatternFill("solid", fgColor="2563EB")
        hdr_font = Font(bold=True, color="FFFFFF", size=12)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        examples = [
            [
                "O'zbekiston qachon mustaqillik oldi?",
                "1990-yil",
                "1991-yil",
                "1992-yil",
                "1993-yil",
                "B",
                "easy",
            ],
            [
                "Dunyo bo'yicha eng katta okean qaysi?",
                "Atlantika okeani",
                "Hind okeani",
                "Tinch okean",
                "Arktika okeani",
                "C",
                "medium",
            ],
            [
                "2 darajaga ko'tarilgan 10 qanday son?",
                "20",
                "12",
                "100",
                "1000",
                "C",
                "easy",
            ],
        ]
        ex_font = Font(size=11)
        for r_idx, row in enumerate(examples, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = ex_font

        col_widths = [48, 22, 22, 22, 22, 22, 26]
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = w
        ws.row_dimensions[1].height = 32

        try:
            wb.save(path)
        except Exception as e:
            QMessageBox.critical(self, t("common.error"), f"Fayl saqlanmadi:\n{e}")
            return

        QMessageBox.information(
            self, t("fan.tpl_saved"),
            t("fan.tpl_saved_msg", path=path)
        )

    def _import_excel(self):
        """Exceldan savollarni o'qib API orqali yuklash."""
        if not self._fan_id:
            QMessageBox.information(self, t("fan.no_fan_title"), t("fan.no_fan_hint"))
            return

        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, t("common.error"), "openpyxl kutubxonasi topilmadi!")
            return

        path, _ = QFileDialog.getOpenFileName(
            self, t("sw.excel_open_title"), "",
            "Excel fayllar (*.xlsx *.xls)"
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(self, t("qs.err_open"), str(e))
            return

        valid_rows = []
        error_lines = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            text   = str(row[0]).strip() if row[0] else ""
            opt_a  = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            opt_b  = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            opt_c  = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            opt_d  = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            answer = str(row[5]).strip().upper() if len(row) > 5 and row[5] else ""
            diff   = str(row[6]).strip().lower() if len(row) > 6 and row[6] else "medium"

            if not all([text, opt_a, opt_b, opt_c, opt_d]):
                error_lines.append(f"• Qator {i}: ba'zi maydonlar bo'sh")
                continue
            if answer not in ("A", "B", "C", "D"):
                error_lines.append(
                    f"• Qator {i}: to'g'ri javob '{answer}' noto'g'ri — A/B/C/D bo'lishi kerak"
                )
                continue
            if diff not in ("easy", "medium", "hard"):
                diff = "medium"

            valid_rows.append({
                "text": text,
                "option_a": opt_a,
                "option_b": opt_b,
                "option_c": opt_c,
                "option_d": opt_d,
                "correct_answer": answer,
                "category_id": self._fan_id,
                "difficulty": diff,
            })

        if not valid_rows:
            msg = t("fan.no_q_msg")
            if error_lines:
                msg += t("common.error") + ":\n" + "\n".join(error_lines[:8])
            else:
                msg += "1-qator sarlavha. 2-qatordan boshlab savol kiriting.\n"
                msg += "«📄 Namuna Excel» tugmasi bilan namuna faylni yuklab oling."
            QMessageBox.warning(self, t("fan.no_q_found"), msg)
            return

        confirm_msg = t("fan.import_confirm", n=len(valid_rows), name=self._fan_name)
        if error_lines:
            confirm_msg += t("fan.import_err_rows", n=len(error_lines))
        confirm_msg += t("fan.import_continue")

        if QMessageBox.question(
            self, t("fan.import_confirm_title"), confirm_msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) != QMessageBox.StandardButton.Yes:
            return

        # Progress dialog
        progress = QProgressDialog(
            t("fan.import_loading"), t("fan.import_cancel"),
            0, len(valid_rows), self
        )
        progress.setWindowTitle(t("fan.loading"))
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumWidth(340)

        success_n = 0
        fail_n = 0
        fail_msgs = []

        for idx, q_data in enumerate(valid_rows):
            if progress.wasCanceled():
                break
            progress.setValue(idx)
            progress.setLabelText(
                t("fan.import_progress", cur=idx+1, total=len(valid_rows), text=q_data['text'][:60])
            )
            try:
                api.create_question(q_data)
                success_n += 1
            except APIError as e:
                fail_n += 1
                fail_msgs.append(str(e))

        progress.setValue(len(valid_rows))

        self._load_questions()
        self.refresh()

        result = t("fan.import_ok", n=success_n)
        if fail_n:
            result += t("fan.import_fail", n=fail_n)
        if error_lines:
            result += t("fan.import_skip", n=len(error_lines))
        QMessageBox.information(self, t("fan.import_result"), result)
