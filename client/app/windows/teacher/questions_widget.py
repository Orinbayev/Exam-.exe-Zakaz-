"""
Savollar boshqaruvi widget - CRUD, qidirish, kategoriya.
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QLineEdit, QComboBox,
    QDialog, QFormLayout, QTextEdit, QMessageBox, QHeaderView,
    QDialogButtonBox, QFrame, QFileDialog
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from ...worker import ApiWorker
from PyQt6.QtGui import QFont, QColor
from ...api_client import api, APIError
from ...styles import COLORS
from ...i18n import t
import openpyxl


class QuestionDialog(QDialog):
    def __init__(self, categories: list, question: dict = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(t("q.dialog_add") if not question else t("q.dialog_edit"))
        self.setMinimumWidth(600)
        self.question = question
        self.categories = categories
        self._setup_ui()
        if question:
            self._fill_data(question)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        form = QFormLayout()
        form.setSpacing(10)

        self.text_edit = QTextEdit()
        self.text_edit.setPlaceholderText(t("q.text_ph"))
        self.text_edit.setMaximumHeight(100)
        form.addRow(t("q.text_label"), self.text_edit)

        self.opt_a = QLineEdit()
        self.opt_a.setPlaceholderText("A variant")
        self.opt_b = QLineEdit()
        self.opt_b.setPlaceholderText("B variant")
        self.opt_c = QLineEdit()
        self.opt_c.setPlaceholderText("C variant")
        self.opt_d = QLineEdit()
        self.opt_d.setPlaceholderText("D variant")
        form.addRow("A variant *:", self.opt_a)
        form.addRow("B variant *:", self.opt_b)
        form.addRow("C variant *:", self.opt_c)
        form.addRow("D variant *:", self.opt_d)

        self.correct_combo = QComboBox()
        for ch in ["A", "B", "C", "D"]:
            self.correct_combo.addItem(ch)
        form.addRow(t("q.correct_label"), self.correct_combo)

        self.category_combo = QComboBox()
        self.category_combo.addItem(t("qs.cat_none"), None)
        for cat in self.categories:
            self.category_combo.addItem(cat["name"], cat["id"])
        form.addRow(t("q.fan_label"), self.category_combo)

        self.difficulty_combo = QComboBox()
        for d in [(t("q.diff_easy"), "easy"), (t("q.diff_med"), "medium"), (t("q.diff_hard"), "hard")]:
            self.difficulty_combo.addItem(d[0], d[1])
        self.difficulty_combo.setCurrentIndex(1)
        form.addRow(t("q.diff_label"), self.difficulty_combo)

        layout.addLayout(form)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _fill_data(self, q: dict):
        self.text_edit.setPlainText(q.get("text", ""))
        self.opt_a.setText(q.get("option_a", ""))
        self.opt_b.setText(q.get("option_b", ""))
        self.opt_c.setText(q.get("option_c", ""))
        self.opt_d.setText(q.get("option_d", ""))
        idx = ["A", "B", "C", "D"].index(q.get("correct_answer", "A"))
        self.correct_combo.setCurrentIndex(idx)
        difficulty_map = {"easy": 0, "medium": 1, "hard": 2}
        self.difficulty_combo.setCurrentIndex(difficulty_map.get(q.get("difficulty", "medium"), 1))

    def _validate_and_accept(self):
        if not self.text_edit.toPlainText().strip():
            QMessageBox.warning(self, t("common.error"), t("qs.val_warn_text"))
            return
        for field, name in [(self.opt_a, "A"), (self.opt_b, "B"), (self.opt_c, "C"), (self.opt_d, "D")]:
            if not field.text().strip():
                QMessageBox.warning(self, t("common.error"), t("qs.val_warn_opt", opt=name))
                return
        self.accept()

    def get_data(self) -> dict:
        return {
            "text": self.text_edit.toPlainText().strip(),
            "option_a": self.opt_a.text().strip(),
            "option_b": self.opt_b.text().strip(),
            "option_c": self.opt_c.text().strip(),
            "option_d": self.opt_d.text().strip(),
            "correct_answer": self.correct_combo.currentText(),
            "category_id": self.category_combo.currentData(),
            "difficulty": self.difficulty_combo.currentData(),
        }


class QuestionsWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.questions = []
        self.categories = []
        self._setup_ui()
        self.refresh()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        # Toolbar
        toolbar = QHBoxLayout()
        title = QLabel(t("qs.title_lbl"))
        title.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(t("qs.search_ph"))
        self.search_input.setMaximumWidth(220)
        self.search_input.textChanged.connect(self._apply_filter)

        self.cat_filter = QComboBox()
        self.cat_filter.setMaximumWidth(160)
        self.cat_filter.addItem(t("qs.all_cats"), None)
        self.cat_filter.currentIndexChanged.connect(self.refresh)

        add_btn = QPushButton(t("qs.add_btn"))
        add_btn.setObjectName("success")
        add_btn.clicked.connect(self._add_question)

        template_btn = QPushButton(t("qs.tpl_btn"))
        template_btn.setObjectName("secondary")
        template_btn.setToolTip(t("qs.tpl_btn"))
        template_btn.clicked.connect(self._download_template)

        import_btn = QPushButton(t("qs.import_btn"))
        import_btn.setObjectName("secondary")
        import_btn.clicked.connect(self._import_excel)

        add_cat_btn = QPushButton(t("qs.add_cat_btn"))
        add_cat_btn.setObjectName("secondary")
        add_cat_btn.clicked.connect(self._add_category)

        refresh_btn = QPushButton("🔄")
        refresh_btn.setMaximumWidth(38)
        refresh_btn.setObjectName("secondary")
        refresh_btn.clicked.connect(self.refresh)

        toolbar.addWidget(title)
        toolbar.addStretch()
        toolbar.addWidget(self.search_input)
        toolbar.addWidget(self.cat_filter)
        toolbar.addWidget(add_cat_btn)
        toolbar.addWidget(template_btn)
        toolbar.addWidget(import_btn)
        toolbar.addWidget(add_btn)
        toolbar.addWidget(refresh_btn)
        layout.addLayout(toolbar)

        # Stats bar
        self.stats_label = QLabel("")
        self.stats_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 12px;")
        layout.addWidget(self.stats_label)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(
            [t("qs.col_id"), t("qs.col_text"), t("qs.col_cat"), t("qs.col_diff"),
             t("qs.col_correct"), t("qs.col_date"), t("qs.col_actions")]
        )
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(2, 130)
        self.table.setColumnWidth(3, 90)
        self.table.setColumnWidth(4, 100)
        self.table.setColumnWidth(5, 130)
        self.table.setColumnWidth(6, 110)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet("""
            QTableWidget { alternate-background-color: rgba(255,255,255,0.03); }
        """)
        layout.addWidget(self.table)

    def refresh(self):
        self.stats_label.setText(t("qs.loading"))
        self._w_cats = ApiWorker(api.get_categories)
        self._w_cats.done.connect(self._apply_cats)
        self._w_cats.error.connect(lambda e: self.stats_label.setText(f"{t('common.error')}: {e}"))
        self._w_cats.start()

    def _apply_cats(self, categories):
        self.categories = categories
        current_cat_id = self.cat_filter.currentData()
        self.cat_filter.blockSignals(True)
        self.cat_filter.clear()
        self.cat_filter.addItem(t("qs.all_cats"), None)
        for cat in self.categories:
            self.cat_filter.addItem(cat["name"], cat["id"])
        self.cat_filter.blockSignals(False)

        cat_id = self.cat_filter.currentData()
        self._w_qs = ApiWorker(api.get_questions, category_id=cat_id)
        self._w_qs.done.connect(self._apply_questions)
        self._w_qs.error.connect(lambda e: self.stats_label.setText(f"{t('common.error')}: {e}"))
        self._w_qs.start()

    def _apply_questions(self, questions):
        self.questions = questions
        self._render_table(self.questions)
        self.stats_label.setText(t("qs.total", n=len(self.questions)))

    def _apply_filter(self):
        search = self.search_input.text().lower()
        filtered = [q for q in self.questions
                    if search in q.get("text", "").lower()]
        self._render_table(filtered)

    def _render_table(self, questions: list):
        self.table.setRowCount(len(questions))
        difficulty_labels = {"easy": t("qs.diff_easy"), "medium": t("qs.diff_med"), "hard": t("qs.diff_hard")}
        cat_map = {c["id"]: c["name"] for c in self.categories}

        for row, q in enumerate(questions):
            self.table.setRowHeight(row, 48)
            self.table.setItem(row, 0, QTableWidgetItem(str(q["id"])))

            text_item = QTableWidgetItem(q["text"][:80] + ("..." if len(q["text"]) > 80 else ""))
            self.table.setItem(row, 1, text_item)

            cat_name = cat_map.get(q.get("category_id"), "—")
            self.table.setItem(row, 2, QTableWidgetItem(cat_name))

            diff_item = QTableWidgetItem(difficulty_labels.get(q.get("difficulty", "medium"), ""))
            self.table.setItem(row, 3, diff_item)

            ans_map = {"A": "🅐 A", "B": "🅑 B", "C": "🅒 C", "D": "🅓 D"}
            ans_item = QTableWidgetItem(ans_map.get(q.get("correct_answer", ""), q.get("correct_answer", "")))
            ans_item.setForeground(QColor(COLORS["success_light"]))
            self.table.setItem(row, 4, ans_item)

            date_str = q.get("created_at", "")[:10] if q.get("created_at") else "—"
            self.table.setItem(row, 5, QTableWidgetItem(date_str))

            # Amallar
            btn_widget = QWidget()
            btn_widget.setStyleSheet("background: transparent;")
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(6, 5, 6, 5)
            btn_layout.setSpacing(6)
            btn_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

            edit_btn = QPushButton("✏️")
            edit_btn.setFixedSize(38, 32)
            edit_btn.setObjectName("table_action")
            edit_btn.setToolTip(t("qs.tip_edit"))
            edit_btn.clicked.connect(lambda _, qid=q["id"]: self._edit_question(qid))

            del_btn = QPushButton("🗑️")
            del_btn.setFixedSize(38, 32)
            del_btn.setObjectName("table_action_danger")
            del_btn.setToolTip(t("qs.tip_del"))
            del_btn.clicked.connect(lambda _, qid=q["id"]: self._delete_question(qid))

            btn_layout.addWidget(edit_btn)
            btn_layout.addWidget(del_btn)
            self.table.setCellWidget(row, 6, btn_widget)

    def _add_question(self):
        dlg = QuestionDialog(self.categories, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            try:
                api.create_question(dlg.get_data())
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, t("common.error"), str(e))

    def _edit_question(self, q_id: int):
        question = next((q for q in self.questions if q["id"] == q_id), None)
        if not question:
            return
        dlg = QuestionDialog(self.categories, question=question, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            try:
                api.update_question(q_id, dlg.get_data())
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, t("common.error"), str(e))

    def _delete_question(self, q_id: int):
        reply = QMessageBox.question(self, t("qs.del_title"),
                                     t("qs.del_confirm"),
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                api.delete_question(q_id)
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, t("common.error"), str(e))

    def _add_category(self):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self, t("qs.cat_add_title"), t("qs.cat_add_prompt"))
        if ok and name.strip():
            try:
                api.create_category(name.strip())
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, t("common.error"), str(e))

    def _download_template(self):
        """Savollar import uchun namuna Excel faylni yaratib saqlash."""
        from PyQt6.QtWidgets import QFileDialog
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        save_path, _ = QFileDialog.getSaveFileName(
            self, t("qs.tpl_title"),
            t("qs.tpl_ph"), "Excel (*.xlsx)"
        )
        if not save_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Savollar"

        BLUE_FILL   = PatternFill("solid", fgColor="1565C0")
        YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")
        HDR_FONT    = Font(bold=True, color="FFFFFF", size=11)
        THIN        = Border(*(Side(style="thin"),)*4)

        headers = [
            ("Savol matni (to'liq)",         45),
            ("A variant",                     22),
            ("B variant",                     22),
            ("C variant",                     22),
            ("D variant",                     22),
            ("To'g'ri javob (A/B/C/D)",       18),
            ("Qiyinlik (easy/medium/hard)",   20),
        ]
        for col, (h, w) in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = HDR_FONT
            cell.fill = BLUE_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 28

        samples = [
            ["O'zbekiston qachon mustaqil bo'ldi?",
             "1990-yil", "1991-yil", "1992-yil", "1993-yil", "B", "easy"],
            ["Quyoshgacha bo'lgan masofa taxminan qancha?",
             "100 mln km", "150 mln km", "200 mln km", "50 mln km", "B", "medium"],
            ["Pythonda ro'yxat (list) qanday belgilanadi?",
             "()", "{}","[]","<>","C","easy"],
            ["Eng katta sayyora qaysi?",
             "Zuhro","Saturn","Yupiter","Mars","C","medium"],
            ["DNA nimaning qisqartmasi?",
             "Deoksiribonuklein kislota","Ribonuklein kislota",
             "Dezoksiriboza","Nuklein asosi","A","hard"],
        ]
        for r_idx, row in enumerate(samples, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.fill = YELLOW_FILL
                cell.border = THIN
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[r_idx].height = 26

        # Ko'rsatma
        ws.cell(len(samples)+3, 1).value = (
            "ESLATMA: 1-qator sarlavha (o'zgartirmang). "
            "2-qatordan boshlab savollarni kiriting. "
            "To'g'ri javob: A, B, C yoki D. "
            "Qiyinlik: easy, medium yoki hard."
        )
        ws.cell(len(samples)+3, 1).font = Font(italic=True, color="555555")
        ws.freeze_panes = "A2"

        try:
            wb.save(save_path)
            QMessageBox.information(
                self, t("qs.tpl_saved_title"),
                t("qs.tpl_saved_msg", path=save_path)
            )
            import os
            if os.name == "nt":
                os.startfile(save_path)
        except Exception as e:
            QMessageBox.critical(self, t("qs.err_save"), t("qs.err_save_msg", e=e))

    def _import_excel(self):
        """Excel fayldan savollar yuklash."""
        file_path, _ = QFileDialog.getOpenFileName(
            self, t("qs.import_title"), "", "Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            imported = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                try:
                    data = {
                        "text": str(row[0]).strip(),
                        "option_a": str(row[1]).strip() if row[1] else "",
                        "option_b": str(row[2]).strip() if row[2] else "",
                        "option_c": str(row[3]).strip() if row[3] else "",
                        "option_d": str(row[4]).strip() if row[4] else "",
                        "correct_answer": str(row[5]).strip().upper() if row[5] else "A",
                        "difficulty": str(row[6]).strip().lower() if len(row) > 6 and row[6] else "medium",
                    }
                    if data["correct_answer"] not in ("A", "B", "C", "D"):
                        data["correct_answer"] = "A"
                    if data["difficulty"] not in ("easy", "medium", "hard"):
                        data["difficulty"] = "medium"
                    api.create_question(data)
                    imported += 1
                except Exception as e:
                    errors.append(f"Qator {row_idx}: {e}")

            msg = t("qs.import_ok", n=imported)
            if errors:
                msg += t("qs.import_err", n=len(errors)) + "\n".join(errors[:5])
            QMessageBox.information(self, t("qs.import_result"), msg)
            self.refresh()
        except Exception as e:
            QMessageBox.critical(self, t("common.error"), f"Fayl ochishda xato: {e}")
