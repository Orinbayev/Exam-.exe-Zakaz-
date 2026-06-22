"""
Sinflar va O'quvchilar boshqaruvi — professional dizayn.
"""
import os
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QListWidget, QListWidgetItem, QTableWidget, QTableWidgetItem,
    QDialog, QCheckBox,
    QLineEdit, QMessageBox,
    QHeaderView, QFrame, QFileDialog,
)
from PyQt6.QtCore import Qt, QSize, QTimer
from ...worker import ApiWorker
from PyQt6.QtGui import QFont, QColor
from ...api_client import api, APIError
from ...styles import COLORS


# ─────────────────────────────────────────────────────────────────────────────
# Yordamchi: zamonaviy dialog bazasi
# ─────────────────────────────────────────────────────────────────────────────

_DLG_STYLE = f"""
QDialog {{
    background-color: {COLORS['bg_medium']};
    color: white;
    font-family: 'Segoe UI', Arial;
}}
QLabel  {{ color: white; background: transparent; }}
QLineEdit {{
    background: {COLORS['bg_dark']};
    color: white;
    border: 1.5px solid {COLORS['border']};
    border-radius: 8px;
    padding: 8px 12px;
    font-size: 13px;
    min-height: 36px;
}}
QLineEdit:focus {{
    border: 2px solid {COLORS['primary_light']};
}}
QPushButton {{
    background: {COLORS['primary']};
    color: white;
    border: none;
    border-radius: 8px;
    padding: 9px 22px;
    font-size: 13px;
    font-weight: bold;
    min-height: 36px;
}}
QPushButton:hover  {{ background: {COLORS['primary_light']}; }}
QPushButton#cancel {{
    background: {COLORS['bg_light']};
    border: 1px solid {COLORS['border']};
}}
QPushButton#cancel:hover {{ background: {COLORS['bg_dark']}; }}
"""


# ─────────────────────────────────────────────────────────────────────────────
# Sinf qo'shish dialogi
# ─────────────────────────────────────────────────────────────────────────────

class ClassDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Yangi sinf qo'shish")
        self.setFixedSize(440, 290)
        self.setStyleSheet(f"""
            QDialog {{
                background: {COLORS['bg_medium']};
                border-radius: 16px;
            }}
        """)
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Header ─────────────────────────────────────────────────────────────
        header = QFrame()
        header.setFixedHeight(100)
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1A237E, stop:1 {COLORS['primary']});
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
            }}
        """)
        h_lay = QVBoxLayout(header)
        h_lay.setContentsMargins(0, 14, 0, 14)
        h_lay.setSpacing(4)

        ico = QLabel("🏫")
        ico.setFont(QFont("Segoe UI Emoji", 26))
        ico.setAlignment(Qt.AlignmentFlag.AlignCenter)
        h_lay.addWidget(ico)

        title_lbl = QLabel("Yangi sinf qo'shish")
        title_lbl.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_lbl.setStyleSheet("color: white; background: transparent;")
        h_lay.addWidget(title_lbl)
        root.addWidget(header)

        # ── Body ───────────────────────────────────────────────────────────────
        body = QFrame()
        body.setStyleSheet(f"""
            QFrame {{
                background: {COLORS['bg_medium']};
                border-bottom-left-radius: 10px;
                border-bottom-right-radius: 10px;
            }}
        """)
        b_lay = QVBoxLayout(body)
        b_lay.setContentsMargins(28, 20, 28, 24)
        b_lay.setSpacing(14)

        lbl = QLabel("Sinf nomini kiriting:")
        lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
        lbl.setStyleSheet(f"color: {COLORS['text_secondary']}; background: transparent;")
        b_lay.addWidget(lbl)

        self.inp = QLineEdit()
        self.inp.setPlaceholderText("Masalan:  9-A   yoki   11-B")
        self.inp.setFixedHeight(46)
        self.inp.setFont(QFont("Segoe UI", 12))
        self.inp.setStyleSheet(f"""
            QLineEdit {{
                background: {COLORS['bg_dark']};
                color: white;
                border: 2px solid {COLORS['border']};
                border-radius: 10px;
                padding: 0 14px;
                font-size: 13px;
            }}
            QLineEdit:focus {{
                border: 2px solid {COLORS['primary_light']};
                background: #0D2137;
            }}
        """)
        self.inp.returnPressed.connect(self._ok)
        b_lay.addWidget(self.inp)

        b_lay.addSpacing(4)

        # ── Buttons ────────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)

        cl_btn = QPushButton("✕   Bekor")
        cl_btn.setFixedSize(140, 42)
        cl_btn.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        cl_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                color: {COLORS['text_secondary']};
                border: 2px solid {COLORS['border']};
                border-radius: 10px;
                font-size: 12px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: {COLORS['bg_light']};
                color: white;
                border-color: {COLORS['primary_light']};
            }}
        """)
        cl_btn.clicked.connect(self.reject)

        ok_btn = QPushButton("✅   Saqlash")
        ok_btn.setFixedSize(160, 42)
        ok_btn.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        ok_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {COLORS['success']}, stop:1 #388E3C);
                color: white;
                border: none;
                border-radius: 10px;
                font-size: 12px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #388E3C, stop:1 {COLORS['success_light']});
            }}
            QPushButton:pressed {{
                background: {COLORS['success']};
            }}
        """)
        ok_btn.clicked.connect(self._ok)

        btn_row.addStretch()
        btn_row.addWidget(cl_btn)
        btn_row.addWidget(ok_btn)
        btn_row.addStretch()
        b_lay.addLayout(btn_row)

        root.addWidget(body)

    def _ok(self):
        if not self.inp.text().strip():
            self.inp.setPlaceholderText("⚠  Sinf nomini kiriting!")
            self.inp.setStyleSheet(self.inp.styleSheet() +
                                   "border: 2px solid #EF5350 !important;")
            return
        self.accept()

    def get_name(self): return self.inp.text().strip()


# ─────────────────────────────────────────────────────────────────────────────
# O'quvchi qo'shish / tahrirlash dialogi
# ─────────────────────────────────────────────────────────────────────────────

class StudentDialog(QDialog):
    def __init__(self, student=None, parent=None):
        super().__init__(parent)
        self.student = student
        self.setWindowTitle(
            "O'quvchi qo'shish" if not student else "O'quvchini tahrirlash"
        )
        self.setFixedWidth(460)
        self.setStyleSheet(_DLG_STYLE)
        self._build()
        if student:
            self.l_inp.setText(student.get("last_name", ""))
            self.f_inp.setText(student.get("first_name", ""))
            self.tg_inp.setText(student.get("parent_telegram_id") or "")

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(28, 22, 28, 22)
        lay.setSpacing(0)

        # Icon
        ico = QLabel("🎒" if not self.student else "✏️")
        ico.setFont(QFont("Segoe UI Emoji", 28))
        ico.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(ico)
        lay.addSpacing(16)

        # ── Fields ──────────────────────────────────────────────────────────
        self.l_inp = self._field(lay, "Familiya *", "O'quvchi familiyasi")
        self.f_inp = self._field(lay, "Ism *", "O'quvchi ismi")
        self.tg_inp = self._field(
            lay,
            "Ota-ona Telegram Chat ID  (ixtiyoriy)",
            "Masalan: 123456789",
            last=True,
        )

        # ── Info card ───────────────────────────────────────────────────────
        info = QFrame()
        info.setStyleSheet("""
            QFrame {
                background: rgba(66,165,245,0.10);
                border: 1px solid rgba(66,165,245,0.28);
                border-radius: 8px;
            }
        """)
        il = QHBoxLayout(info)
        il.setContentsMargins(12, 9, 12, 9)
        il.setSpacing(8)
        info_ico = QLabel("ℹ️")
        info_ico.setFont(QFont("Segoe UI Emoji", 13))
        info_ico.setStyleSheet("background: transparent;")
        info_ico.setFixedWidth(22)
        info_txt = QLabel(
            "Chat ID ni topish: Telegramda  <b>@userinfobot</b>  ga  /start  yuboring"
        )
        info_txt.setFont(QFont("Segoe UI", 10))
        info_txt.setStyleSheet(f"color: {COLORS['primary_light']}; background: transparent;")
        info_txt.setWordWrap(True)
        il.addWidget(info_ico)
        il.addWidget(info_txt, stretch=1)
        lay.addWidget(info)
        lay.addSpacing(20)

        # ── Buttons ─────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        cancel_btn = QPushButton("Bekor qilish")
        cancel_btn.setObjectName("cancel")
        cancel_btn.clicked.connect(self.reject)
        save_btn = QPushButton("✅  Saqlash")
        save_btn.clicked.connect(self._ok)
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

    @staticmethod
    def _field(layout: QVBoxLayout, label: str, placeholder: str,
               last: bool = False) -> QLineEdit:
        """Label + input qo'shuvchi yordamchi — to'g'ri oraliq bilan."""
        lbl = QLabel(label)
        lbl.setFont(QFont("Segoe UI", 10, QFont.Weight.DemiBold))
        lbl.setStyleSheet(f"color: {COLORS['text_secondary']}; background: transparent;")
        inp = QLineEdit()
        inp.setPlaceholderText(placeholder)
        inp.setFixedHeight(42)
        layout.addWidget(lbl)
        layout.addSpacing(5)
        layout.addWidget(inp)
        layout.addSpacing(0 if last else 16)
        return inp

    # ── Validation ────────────────────────────────────────────────────────────

    def _ok(self):
        if not self.l_inp.text().strip():
            self._shake(self.l_inp, "Familiya bo'sh bo'lishi mumkin emas!")
            return
        if not self.f_inp.text().strip():
            self._shake(self.f_inp, "Ism bo'sh bo'lishi mumkin emas!")
            return
        self.accept()

    def _shake(self, inp: QLineEdit, msg: str):
        orig = inp.styleSheet()
        inp.setStyleSheet(orig + "border: 2px solid #EF5350;")
        QMessageBox.warning(self, "⚠  Xato", msg)
        inp.setStyleSheet(orig)
        inp.setFocus()

    def get_data(self) -> dict:
        return {
            "first_name": self.f_inp.text().strip(),
            "last_name":  self.l_inp.text().strip(),
            "parent_telegram_id": self.tg_inp.text().strip() or None,
        }


# ─────────────────────────────────────────────────────────────────────────────
# Excel format ko'rsatish dialogi
# ─────────────────────────────────────────────────────────────────────────────

class ExcelFormatDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Excel import formati")
        self.setFixedSize(580, 460)
        self.setStyleSheet(_DLG_STYLE)
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(28, 20, 28, 20)
        lay.setSpacing(14)

        title = QLabel("📊  O'quvchilar Excel import formati")
        title.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(title)

        # Ustunlar tavsifi
        cols_card = QFrame()
        cols_card.setStyleSheet(f"""
            QFrame {{
                background: {COLORS['bg_dark']};
                border: 1px solid {COLORS['border']};
                border-radius: 10px;
            }}
        """)
        clay = QVBoxLayout(cols_card)
        clay.setContentsMargins(16, 14, 16, 14)
        clay.setSpacing(8)

        hdr = QLabel("Ustunlar tartibi:")
        hdr.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        hdr.setStyleSheet(f"color: {COLORS['primary_light']};")
        clay.addWidget(hdr)

        cols_info = [
            ("A ustun", "Ism",                       "Majburiy",   "#66BB6A"),
            ("B ustun", "Familiya",                   "Majburiy",   "#66BB6A"),
            ("C ustun", "Ota-ona Telegram Chat ID",   "Ixtiyoriy",  "#FFA726"),
        ]
        for col, desc, req, color in cols_info:
            row_w = QWidget()
            row_w.setStyleSheet("background: transparent;")
            rlay = QHBoxLayout(row_w)
            rlay.setContentsMargins(0, 0, 0, 0)
            rlay.setSpacing(10)

            col_lbl = QLabel(col)
            col_lbl.setFixedWidth(72)
            col_lbl.setFont(QFont("Consolas", 11, QFont.Weight.Bold))
            col_lbl.setStyleSheet(f"color: #FFD740; background: transparent;")

            desc_lbl = QLabel(desc)
            desc_lbl.setFont(QFont("Segoe UI", 11))
            desc_lbl.setStyleSheet("background: transparent;")

            req_lbl = QLabel(req)
            req_lbl.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            req_lbl.setStyleSheet(f"color: {color}; background: transparent;")
            req_lbl.setAlignment(Qt.AlignmentFlag.AlignRight)

            rlay.addWidget(col_lbl)
            rlay.addWidget(desc_lbl, stretch=1)
            rlay.addWidget(req_lbl)
            clay.addWidget(row_w)

        lay.addWidget(cols_card)

        # Namuna jadval
        ex_lbl = QLabel("Namuna ko'rinish:")
        ex_lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        ex_lbl.setStyleSheet(f"color: {COLORS['text_secondary']};")
        lay.addWidget(ex_lbl)

        tbl = QTableWidget(4, 3)
        tbl.setHorizontalHeaderLabels(["A — Ism", "B — Familiya", "C — Telegram ID"])
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        tbl.setFixedHeight(138)
        tbl.setStyleSheet(f"""
            QTableWidget {{
                background: {COLORS['bg_dark']};
                color: white;
                border: 1px solid {COLORS['border']};
                border-radius: 8px;
                gridline-color: {COLORS['border']};
                font-size: 12px;
            }}
            QHeaderView::section {{
                background: {COLORS['bg_light']};
                color: {COLORS['text_secondary']};
                padding: 6px;
                border: none;
                font-size: 11px;
                font-weight: bold;
            }}
            QTableWidget::item {{ padding: 6px 10px; }}
        """)

        sample_data = [
            ["Alisher",  "Navoiy",   "123456789"],
            ["Nodira",   "Karimova", "987654321"],
            ["Jasur",    "Toshmatov",""],
            ["Malika",   "Yusupova", "456123789"],
        ]
        for r, row in enumerate(sample_data):
            for c, val in enumerate(row):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if c == 2 and val:
                    item.setForeground(QColor("#FFA726"))
                tbl.setItem(r, c, item)
        lay.addWidget(tbl)

        # Eslatma
        note = QLabel(
            "⚠️  1-qator sarlavha sifatida o'tkazib yuboriladi.  "
            "C ustun bo'sh bo'lsa ham bo'ladi — Telegram xabar yuborilmaydi."
        )
        note.setFont(QFont("Segoe UI", 10))
        note.setStyleSheet(f"color: {COLORS['accent_light']}; background: transparent;")
        note.setWordWrap(True)
        lay.addWidget(note)

        # Tugmalar
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        dl_btn = QPushButton("📥  Namuna Excel yuklab olish")
        dl_btn.clicked.connect(self._download)
        cl_btn = QPushButton("Yopish")
        cl_btn.setObjectName("cancel")
        cl_btn.clicked.connect(self.accept)
        btn_row.addWidget(dl_btn)
        btn_row.addWidget(cl_btn)
        lay.addLayout(btn_row)

    def _download(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Namuna faylni saqlash",
            "oquvchilar_namuna.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        _generate_student_template(path)
        QMessageBox.information(
            self, "✅ Saqlandi",
            f"Namuna fayl saqlandi:\n{path}\n\n"
            "Sariq qatorlardagi ma'lumotlarni o'chirib,\n"
            "o'z o'quvchilaringizni kiriting."
        )
        if os.name == "nt":
            os.startfile(path)


def _generate_student_template(path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"

    BLU = PatternFill("solid", fgColor="1565C0")
    YEL = PatternFill("solid", fgColor="FFF9C4")
    HDR = Font(bold=True, color="FFFFFF", size=11)
    BRD = Border(*(Side(style="thin"),)*4)
    CTR = Alignment(horizontal="center", vertical="center")

    headers = [("Ism (A)", 20), ("Familiya (B)", 22), ("Ota-ona Telegram ID (C)", 28)]
    for c, (h, w) in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font, cell.fill, cell.alignment, cell.border = HDR, BLU, CTR, BRD
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 26

    samples = [
        ["Alisher",  "Navoiy",    "123456789"],
        ["Nodira",   "Karimova",  "987654321"],
        ["Jasur",    "Toshmatov", ""],
        ["Malika",   "Yusupova",  "456123789"],
        ["Sanjar",   "Ergashev",  ""],
    ]
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.fill, cell.border = YEL, BRD
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22

    ws.cell(len(samples)+3, 1).value = (
        "ESLATMA: 1-qatorni o'zgartirmang (sarlavha). "
        "2-qatordan boshlab o'z o'quvchilaringizni kiriting. "
        "Telegram ID ixtiyoriy — bo'sh qoldirsangiz ham bo'ladi."
    )
    ws.cell(len(samples)+3, 1).font = Font(italic=True, color="888888", size=9)
    ws.freeze_panes = "A2"
    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# Asosiy widget
# ─────────────────────────────────────────────────────────────────────────────

_PANEL_STYLE = f"""
    QFrame {{
        background: {COLORS['bg_medium']};
        border: 1px solid {COLORS['border']};
        border-radius: 12px;
    }}
"""

_LIST_STYLE = f"""
    QListWidget {{
        background: {COLORS['bg_dark']};
        border: 1px solid {COLORS['border']};
        border-radius: 8px;
        color: white;
        font-size: 13px;
        outline: none;
    }}
    QListWidget::item {{
        padding: 10px 14px;
        border-bottom: 1px solid {COLORS['border']};
    }}
    QListWidget::item:selected {{
        background: {COLORS['primary']};
        color: white;
        border-radius: 6px;
    }}
    QListWidget::item:hover:!selected {{
        background: {COLORS['bg_light']};
    }}
"""

_TABLE_STYLE = f"""
    QTableWidget {{
        background: {COLORS['bg_dark']};
        color: white;
        border: 1px solid {COLORS['border']};
        border-radius: 8px;
        gridline-color: {COLORS['border']};
        font-size: 13px;
        outline: none;
    }}
    QTableWidget::item {{
        padding: 8px 12px;
    }}
    QTableWidget::item:selected {{
        background: {COLORS['primary']};
        color: white;
    }}
    QHeaderView::section {{
        background: {COLORS['bg_light']};
        color: {COLORS['text_secondary']};
        padding: 10px 12px;
        border: none;
        border-right: 1px solid {COLORS['border']};
        border-bottom: 1px solid {COLORS['border']};
        font-weight: bold;
        font-size: 12px;
    }}
"""



class StudentsWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.classes   = []
        self.students  = []
        self._cls_id   = None
        self._cls_name = ""
        self._setup_ui()
        self.refresh()

    # ── UI qurilishi ──────────────────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(14)

        # ── Sarlavha ──────────────────────────────────────────────────────────
        hdr = QHBoxLayout()
        title = QLabel("👥  Sinflar va O'quvchilar")
        title.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
        hdr.addWidget(title)
        hdr.addStretch()
        fmt_btn = QPushButton("📊  Excel format ko'rsatish")
        fmt_btn.setObjectName("secondary")
        fmt_btn.clicked.connect(self._show_format)
        hdr.addWidget(fmt_btn)
        root.addLayout(hdr)

        # Qisqacha yo'riqnoma
        hint = QFrame()
        hint.setStyleSheet(f"""
            QFrame {{
                background: rgba(21,101,192,0.15);
                border: 1px solid rgba(66,165,245,0.30);
                border-radius: 8px;
            }}
        """)
        hlay = QHBoxLayout(hint)
        hlay.setContentsMargins(14, 10, 14, 10)
        hint_lbl = QLabel(
            "💡  <b>Qanday ishlaydi:</b>  "
            "1) Chapdan sinf tanlang  →  "
            "2) O'quvchi qo'shing  →  "
            "3) Ota-ona Telegram ID kiriting  →  "
            "Test natijasi avtomatik yuboriladi"
        )
        hint_lbl.setFont(QFont("Segoe UI", 11))
        hint_lbl.setStyleSheet(f"color: {COLORS['primary_light']}; background: transparent;")
        hint_lbl.setWordWrap(True)
        hlay.addWidget(hint_lbl)
        root.addWidget(hint)

        # ── Asosiy panel (chap + o'ng) ────────────────────────────────────────
        panels = QHBoxLayout()
        panels.setSpacing(14)

        # ── CHAP: Sinflar ─────────────────────────────────────────────────────
        left = QFrame()
        left.setStyleSheet(_PANEL_STYLE)
        left.setFixedWidth(240)
        ll = QVBoxLayout(left)
        ll.setContentsMargins(12, 14, 12, 14)
        ll.setSpacing(10)

        l_hdr = QHBoxLayout()
        l_title = QLabel("📚  Sinflar")
        l_title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        l_hdr.addWidget(l_title)
        self.cls_count_lbl = QLabel("0 ta")
        self.cls_count_lbl.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 11px;")
        l_hdr.addStretch()
        l_hdr.addWidget(self.cls_count_lbl)
        ll.addLayout(l_hdr)

        self.class_list = QListWidget()
        self.class_list.setStyleSheet(_LIST_STYLE)
        self.class_list.setAlternatingRowColors(False)
        self.class_list.currentItemChanged.connect(self._on_class_selected)
        ll.addWidget(self.class_list)

        toggle_btn = QPushButton("🟢 Faollashtirish / ⚫ To'xtatish")
        toggle_btn.setObjectName("secondary")
        toggle_btn.clicked.connect(self._toggle_class_active)
        ll.addWidget(toggle_btn)

        add_cls_btn = QPushButton("＋  Sinf qo'shish")
        add_cls_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['success']};
                color: white; border: none; border-radius: 8px;
                padding: 10px; font-size: 13px; font-weight: bold;
            }}
            QPushButton:hover {{ background: {COLORS['success_light']}; }}
        """)
        add_cls_btn.clicked.connect(self._add_class)
        ll.addWidget(add_cls_btn)

        del_cls_btn = QPushButton("🗑  Sinf o'chirish")
        del_cls_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['bg_dark']};
                color: {COLORS['danger_light']};
                border: 1px solid {COLORS['danger']};
                border-radius: 8px; padding: 8px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background: {COLORS['danger']};
                color: white;
            }}
        """)
        del_cls_btn.clicked.connect(self._delete_class)
        ll.addWidget(del_cls_btn)

        panels.addWidget(left)

        # ── O'NG: O'quvchilar ─────────────────────────────────────────────────
        right = QFrame()
        right.setStyleSheet(_PANEL_STYLE)
        rl = QVBoxLayout(right)
        rl.setContentsMargins(14, 14, 14, 14)
        rl.setSpacing(10)

        # O'ng sarlavha
        r_hdr = QHBoxLayout()
        self.students_title = QLabel("👤  O'quvchilar")
        self.students_title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        r_hdr.addWidget(self.students_title)
        r_hdr.addStretch()
        self.st_count_lbl = QLabel("")
        self.st_count_lbl.setStyleSheet(f"color: {COLORS['text_secondary']}; font-size: 11px;")
        r_hdr.addWidget(self.st_count_lbl)
        rl.addLayout(r_hdr)

        # Bo'sh holat labeli
        self.empty_lbl = QLabel(
            "⬅  Chapdan sinf tanlang\nSo'ng o'quvchi qo'shing"
        )
        self.empty_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.empty_lbl.setStyleSheet(f"""
            color: {COLORS['text_secondary']};
            font-size: 14px;
            background: transparent;
        """)

        # Jadval
        self.tbl = QTableWidget()
        self.tbl.setColumnCount(4)
        self.tbl.setHorizontalHeaderLabels(
            ["Familiya", "Ism", "Ota-ona Telegram ID", "Amallar"]
        )
        self.tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.tbl.setColumnWidth(2, 200)
        self.tbl.setColumnWidth(3, 110)
        self.tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setStyleSheet(_TABLE_STYLE + """
            QTableWidget { alternate-background-color: rgba(255,255,255,0.03); }
        """)
        self.tbl.hide()

        rl.addWidget(self.empty_lbl, stretch=1)
        rl.addWidget(self.tbl, stretch=1)

        # Tugmalar qatori
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        self.add_st_btn = QPushButton("＋  O'quvchi qo'shish")
        self.add_st_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['success']};
                color: white; border: none; border-radius: 8px;
                padding: 9px 16px; font-size: 13px; font-weight: bold;
            }}
            QPushButton:hover {{ background: {COLORS['success_light']}; }}
            QPushButton:disabled {{ background: {COLORS['bg_light']}; color: {COLORS['text_secondary']}; }}
        """)
        self.add_st_btn.setEnabled(False)
        self.add_st_btn.clicked.connect(self._add_student)

        self.fmt_xl_btn = QPushButton("📋  Namuna Excel")
        self.fmt_xl_btn.setObjectName("secondary")
        self.fmt_xl_btn.clicked.connect(self._show_format)

        self.imp_xl_btn = QPushButton("📥  Excel import")
        self.imp_xl_btn.setObjectName("secondary")
        self.imp_xl_btn.setEnabled(False)
        self.imp_xl_btn.clicked.connect(self._import_excel)

        btn_row.addWidget(self.add_st_btn)
        btn_row.addStretch()
        btn_row.addWidget(self.fmt_xl_btn)
        btn_row.addWidget(self.imp_xl_btn)
        rl.addLayout(btn_row)

        # Fan biriktirish
        fan_link_frame = QFrame()
        fan_link_frame.setStyleSheet(f"""
            QFrame {{
                background: {COLORS['bg_dark']};
                border: 1px solid {COLORS['border']};
                border-radius: 10px;
            }}
        """)
        tl = QVBoxLayout(fan_link_frame)
        tl.setContentsMargins(12, 10, 12, 10)
        tl.setSpacing(8)

        tl_hdr = QHBoxLayout()
        tl_title = QLabel("📚  Biriktirilgan Fanlar")
        tl_title.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        tl_hdr.addWidget(tl_title)
        tl_hdr.addStretch()

        sync_btn = QPushButton("🔄 Yangilash")
        sync_btn.setToolTip("Yangi savollar qo'shilgan bo'lsa — testni yangilaydi")
        sync_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['bg_light']};
                color: #69F0AE; border: 1px solid rgba(105,240,174,0.40);
                border-radius: 6px; padding: 5px 10px; font-size: 11px; font-weight: bold;
            }}
            QPushButton:hover {{ background: rgba(105,240,174,0.15); }}
        """)
        sync_btn.clicked.connect(self._sync_all_fans)
        tl_hdr.addWidget(sync_btn)

        add_fan_btn = QPushButton("📚 Fan biriktirish")
        add_fan_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['primary']};
                color: white; border: none; border-radius: 6px;
                padding: 6px 12px; font-size: 12px; font-weight: bold;
            }}
            QPushButton:hover {{ background: {COLORS['primary_light']}; }}
        """)
        add_fan_btn.clicked.connect(self._assign_fan_dialog)
        tl_hdr.addWidget(add_fan_btn)
        tl.addLayout(tl_hdr)

        self.linked_fans_list = QListWidget()
        self.linked_fans_list.setMaximumHeight(110)
        self.linked_fans_list.setStyleSheet(_LIST_STYLE)
        tl.addWidget(self.linked_fans_list)
        rl.addWidget(fan_link_frame)

        # ── LIVE NATIJALAR ────────────────────────────────────────────────────
        live_frame = QFrame()
        live_frame.setStyleSheet(f"""
            QFrame {{
                background: {COLORS['bg_dark']};
                border: 1px solid rgba(105,240,174,0.25);
                border-radius: 10px;
            }}
        """)
        live_layout = QVBoxLayout(live_frame)
        live_layout.setContentsMargins(12, 10, 12, 10)
        live_layout.setSpacing(7)

        live_hdr = QHBoxLayout()
        live_title = QLabel("📊  Sinf natijalari  (oxirgi)")
        live_title.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        live_title.setStyleSheet("color: #69F0AE; background: transparent;")
        live_hdr.addWidget(live_title)
        live_hdr.addStretch()

        self._live_refresh_btn = QPushButton("🔄")
        self._live_refresh_btn.setFixedSize(30, 28)
        self._live_refresh_btn.setStyleSheet(f"""
            QPushButton {{
                background: {COLORS['bg_light']};
                color: white; border: 1px solid {COLORS['border']};
                border-radius: 6px; font-size: 12px;
            }}
            QPushButton:hover {{ background: {COLORS['primary']}; }}
        """)
        self._live_refresh_btn.clicked.connect(self._load_live_results)
        live_hdr.addWidget(self._live_refresh_btn)
        live_layout.addLayout(live_hdr)

        self._live_table = QTableWidget()
        self._live_table.setColumnCount(5)
        self._live_table.setHorizontalHeaderLabels(
            ["Vaqt", "O'quvchi", "Fan", "To'g'ri", "Baho"]
        )
        self._live_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self._live_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch
        )
        self._live_table.setColumnWidth(0, 90)
        self._live_table.setColumnWidth(3, 70)
        self._live_table.setColumnWidth(4, 55)
        self._live_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._live_table.verticalHeader().setVisible(False)
        self._live_table.setMaximumHeight(160)
        self._live_table.setStyleSheet(_TABLE_STYLE)
        live_layout.addWidget(self._live_table)

        rl.addWidget(live_frame)

        # Auto-refresh every 30s
        self._live_timer = QTimer()
        self._live_timer.timeout.connect(self._load_live_results)
        self._live_timer.start(30_000)

        panels.addWidget(right, stretch=1)
        root.addLayout(panels, stretch=1)

    # ── Ma'lumot yuklash ──────────────────────────────────────────────────────

    def refresh(self):
        self.cls_count_lbl.setText("Yuklanmoqda…")
        self._w_cls = ApiWorker(api.get_classes)
        self._w_cls.done.connect(self._apply_classes)
        self._w_cls.error.connect(lambda _: self._apply_classes([]))
        self._w_cls.start()

    def _apply_classes(self, classes):
        self.classes = classes
        self.class_list.blockSignals(True)
        self.class_list.clear()
        for c in self.classes:
            icon = "🟢" if c.get("is_active") else "⚫"
            item = QListWidgetItem(f" {icon}  {c['name']}")
            item.setSizeHint(QSize(0, 44))
            item.setData(Qt.ItemDataRole.UserRole, c["id"])
            item.setData(Qt.ItemDataRole.UserRole + 1, c.get("student_count", 0))
            self.class_list.addItem(item)
        self.class_list.blockSignals(False)
        self.cls_count_lbl.setText(f"{len(self.classes)} ta")

        if self._cls_id:
            for i in range(self.class_list.count()):
                if self.class_list.item(i).data(Qt.ItemDataRole.UserRole) == self._cls_id:
                    self.class_list.setCurrentRow(i)
                    break

    def _on_class_selected(self, item):
        if not item:
            return
        self._cls_id   = item.data(Qt.ItemDataRole.UserRole)
        self._cls_name = item.text().strip()
        self.students_title.setText(f"👤  {self._cls_name} — O'quvchilar")
        self.add_st_btn.setEnabled(True)
        self.imp_xl_btn.setEnabled(True)
        self._load_students()
        self._load_linked_fans()
        self._load_live_results()

    def _load_students(self):
        if not self._cls_id:
            return
        try:
            self.students = api.get_students(self._cls_id)
        except APIError as e:
            QMessageBox.warning(self, "Xato", str(e))
            return
        self._render()

    def _render(self):
        if not self.students:
            self.tbl.hide()
            self.empty_lbl.setText(
                f"📭  {self._cls_name} sinfida hali o'quvchi yo'q\n"
                "«＋ O'quvchi qo'shish» tugmasini bosing"
            )
            self.empty_lbl.show()
            self.st_count_lbl.setText("")
            return

        self.empty_lbl.hide()
        self.tbl.show()
        self.st_count_lbl.setText(f"{len(self.students)} ta o'quvchi")
        self.tbl.setRowCount(len(self.students))

        for row, s in enumerate(self.students):
            self.tbl.setRowHeight(row, 44)

            for col, val in enumerate([s["last_name"], s["first_name"]]):
                item = QTableWidgetItem(val)
                item.setFont(QFont("Segoe UI", 12))
                self.tbl.setItem(row, col, item)

            tg = s.get("parent_telegram_id") or ""
            tg_item = QTableWidgetItem("✅  " + tg if tg else "—  (kiritilmagan)")
            tg_item.setForeground(
                QColor(COLORS["success_light"]) if tg
                else QColor(COLORS["text_secondary"])
            )
            tg_item.setFont(QFont("Segoe UI", 11))
            self.tbl.setItem(row, 2, tg_item)

            # Amallar
            cell_w = QWidget()
            cell_w.setStyleSheet("background: transparent;")
            cly = QHBoxLayout(cell_w)
            cly.setContentsMargins(6, 4, 6, 4)
            cly.setSpacing(6)
            cly.setAlignment(Qt.AlignmentFlag.AlignCenter)

            ed = QPushButton("✏️")
            ed.setFixedSize(38, 32)
            ed.setObjectName("table_action")
            ed.setToolTip("Tahrirlash")
            ed.clicked.connect(lambda _, sid=s["id"]: self._edit_student(sid))

            dl = QPushButton("🗑️")
            dl.setFixedSize(38, 32)
            dl.setObjectName("table_action_danger")
            dl.setToolTip("O'chirish")
            dl.clicked.connect(lambda _, sid=s["id"]: self._delete_student(sid))

            cly.addWidget(ed)
            cly.addWidget(dl)
            self.tbl.setCellWidget(row, 3, cell_w)

    # ── Sinf CRUD ──────────────────────────────────────────────────────────────

    def _add_class(self):
        dlg = ClassDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            try:
                api.create_class(dlg.get_name())
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, "Xato", str(e))

    def _delete_class(self):
        item = self.class_list.currentItem()
        if not item:
            QMessageBox.information(self, "Tanlash", "Avval chapdan sinf tanlang!")
            return
        name = item.text().strip()
        n = item.data(Qt.ItemDataRole.UserRole + 1)
        msg = (f"«{name}» sinfini o'chirishni tasdiqlaysizmi?"
               + (f"\n\nDiqqat: Bu sinfda {n} ta o'quvchi bor, ular ham o'chiriladi!" if n else ""))
        r = QMessageBox.question(self, "O'chirish", msg,
                                 QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if r == QMessageBox.StandardButton.Yes:
            try:
                api.delete_class(item.data(Qt.ItemDataRole.UserRole))
                self._cls_id = None
                self._cls_name = ""
                self.tbl.hide()
                self.empty_lbl.setText("⬅  Chapdan sinf tanlang\nSo'ng o'quvchi qo'shing")
                self.empty_lbl.show()
                self.students_title.setText("👤  O'quvchilar")
                self.add_st_btn.setEnabled(False)
                self.imp_xl_btn.setEnabled(False)
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, "Xato", str(e))

    # ── O'quvchi CRUD ──────────────────────────────────────────────────────────

    def _add_student(self):
        dlg = StudentDialog(parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            try:
                api.add_student(self._cls_id, dlg.get_data())
                self._load_students()
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, "Xato", str(e))

    def _edit_student(self, sid):
        s = next((x for x in self.students if x["id"] == sid), None)
        if not s:
            return
        dlg = StudentDialog(student=s, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            try:
                api.update_student(sid, dlg.get_data())
                self._load_students()
            except APIError as e:
                QMessageBox.critical(self, "Xato", str(e))

    def _delete_student(self, sid):
        s = next((x for x in self.students if x["id"] == sid), None)
        if not s:
            return
        r = QMessageBox.question(
            self, "O'chirish",
            f"{s['last_name']} {s['first_name']}ni ro'yxatdan o'chirishni tasdiqlaysizmi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if r == QMessageBox.StandardButton.Yes:
            try:
                api.delete_student(sid)
                self._load_students()
                self.refresh()
            except APIError as e:
                QMessageBox.critical(self, "Xato", str(e))

    # ── Sinf aktivlashtirish ──────────────────────────────────────────────────

    def _toggle_class_active(self):
        item = self.class_list.currentItem()
        if not item:
            QMessageBox.information(self, "Tanlash", "Avval chapdan sinf tanlang!")
            return
        cls_id = item.data(Qt.ItemDataRole.UserRole)
        try:
            api.toggle_class_active(cls_id)
            self.refresh()
        except APIError as e:
            QMessageBox.critical(self, "Xato", str(e))

    # ── Fan biriktirish ───────────────────────────────────────────────────────

    def _load_linked_fans(self):
        self.linked_fans_list.clear()
        if not self._cls_id:
            return
        try:
            fans = api.get_class_fans(self._cls_id)
            for f in fans:
                time_lbl = f.get("time_limit", 30)
                item = QListWidgetItem(
                    f"📚  {f['fan_name']}  ·  ⏱ {time_lbl} daqiqa"
                )
                item.setData(Qt.ItemDataRole.UserRole, f["fan_id"])
                item.setForeground(QColor(COLORS["success_light"]))
                self.linked_fans_list.addItem(item)
            if not fans:
                item = QListWidgetItem("— Bu sinfga hali fan biriktirilmagan —")
                item.setForeground(QColor(COLORS["text_secondary"]))
                self.linked_fans_list.addItem(item)
        except APIError:
            pass

    def _assign_fan_dialog(self):
        if not self._cls_id:
            QMessageBox.information(self, "Tanlash", "Avval sinf tanlang!")
            return
        try:
            all_fans = api.get_categories()
            if not all_fans:
                QMessageBox.information(self, "Fan yo'q",
                    "Hali fan yaratilmagan!\n\n«Fan» bo'limidan avval fan yarating.")
                return

            linked_ids = {f["fan_id"] for f in api.get_class_fans(self._cls_id)}

            dlg = QDialog(self)
            dlg.setWindowTitle(f"Fan biriktirish — {self._cls_name}")
            dlg.setMinimumWidth(420)
            dlg.setStyleSheet(f"""
                QDialog {{ background: {COLORS['bg_medium']}; }}
                QLabel {{ color: white; background: transparent; }}
                QCheckBox {{ color: white; font-size: 13px; padding: 6px 0; }}
                QCheckBox::indicator {{ width:18px; height:18px; border-radius:4px;
                    border:2px solid {COLORS['border']}; background:{COLORS['bg_dark']}; }}
                QCheckBox::indicator:checked {{
                    background: {COLORS['primary']}; border-color: {COLORS['primary_light']};
                }}
                QPushButton {{
                    background: {COLORS['primary']}; color: white; border: none;
                    border-radius: 8px; padding: 9px 20px; font-size: 13px; font-weight: bold;
                }}
                QPushButton:hover {{ background: {COLORS['primary_light']}; }}
                QPushButton#cancel {{
                    background: {COLORS['bg_light']}; border: 1px solid {COLORS['border']};
                }}
            """)

            dl = QVBoxLayout(dlg)
            dl.setContentsMargins(24, 20, 24, 20)
            dl.setSpacing(10)

            hdr_lbl = QLabel(f"📚  «{self._cls_name}» sinfi uchun fan tanlang:")
            hdr_lbl.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
            dl.addWidget(hdr_lbl)

            hint = QLabel("✅ Biriktirilgan (saqlasangiz test yangilanadi)  |  ☐ Biriktirilmagan")
            hint.setFont(QFont("Segoe UI", 10))
            hint.setStyleSheet(f"color: {COLORS['text_secondary']};")
            dl.addWidget(hint)

            sep = QFrame()
            sep.setFrameShape(QFrame.Shape.HLine)
            sep.setStyleSheet(f"color: {COLORS['border']};")
            dl.addWidget(sep)

            checkboxes = []
            for f in all_fans:
                cb = QCheckBox(
                    f"  {f['name']}  ·  {f.get('question_count', 0)} ta savol"
                )
                cb.setChecked(f["id"] in linked_ids)
                cb.setProperty("fan_id", f["id"])
                cb.setProperty("fan_name", f["name"])
                dl.addWidget(cb)
                checkboxes.append(cb)

            dl.addSpacing(8)
            info = QLabel(
                "ℹ️  «Saqlash» bosilganda belgilangan fanlar yangilanadi — "
                "yangi qo'shilgan savollar avtomatik testga qo'shiladi."
            )
            info.setFont(QFont("Segoe UI", 10))
            info.setStyleSheet(f"color: {COLORS['text_secondary']};")
            info.setWordWrap(True)
            dl.addWidget(info)

            btn_row = QHBoxLayout()
            btn_row.setSpacing(10)
            cancel_btn = QPushButton("Bekor qilish")
            cancel_btn.setObjectName("cancel")
            cancel_btn.clicked.connect(dlg.reject)
            save_btn = QPushButton("✅  Saqlash")
            save_btn.clicked.connect(dlg.accept)
            btn_row.addWidget(cancel_btn)
            btn_row.addWidget(save_btn)
            dl.addLayout(btn_row)

            if dlg.exec() != QDialog.DialogCode.Accepted:
                return

            errors = []
            for cb in checkboxes:
                fid = cb.property("fan_id")
                fname = cb.property("fan_name")
                was_linked = fid in linked_ids
                now_checked = cb.isChecked()
                if now_checked:
                    # Har doim assign — server yangi savollarni ham qo'shadi
                    try:
                        api.assign_fan_to_class(self._cls_id, fid)
                    except APIError as e:
                        errors.append(f"«{fname}»: {e}")
                elif was_linked and not now_checked:
                    try:
                        api.unassign_fan_from_class(self._cls_id, fid)
                    except APIError as e:
                        errors.append(f"«{fname}» (ajratish): {e}")

            if errors:
                QMessageBox.warning(self, "Ba'zi xatolar", "\n".join(errors))
            self._load_linked_fans()

        except APIError as e:
            QMessageBox.critical(self, "Xato", str(e))

    # ── Fan testlarini yangilash ──────────────────────────────────────────────

    def _sync_all_fans(self):
        """Biriktirilgan barcha fanlarning testini yangi savollar bilan yangilaydi."""
        if not self._cls_id:
            QMessageBox.information(self, "Tanlash", "Avval sinf tanlang!")
            return
        try:
            fans = api.get_class_fans(self._cls_id)
            if not fans:
                QMessageBox.information(self, "Fan yo'q",
                    "Bu sinfga biriktirilgan fan yo'q.")
                return
            updated = []
            errors  = []
            for f in fans:
                try:
                    res = api.assign_fan_to_class(self._cls_id, f["fan_id"])
                    qc = res.get("question_count", "?")
                    updated.append(f"✅  {f['fan_name']}  →  {qc} ta savol")
                except APIError as e:
                    errors.append(f"❌  {f['fan_name']}: {e}")
            msg = "\n".join(updated)
            if errors:
                msg += "\n\n" + "\n".join(errors)
            QMessageBox.information(self, "✅ Yangilandi", msg)
            self._load_linked_fans()
        except APIError as e:
            QMessageBox.critical(self, "Xato", str(e))

    # ── Live natijalar ────────────────────────────────────────────────────────

    def _load_live_results(self):
        if not self._cls_id:
            return
        try:
            all_results = api.get_results()
        except APIError:
            return

        cls_name_clean = self._cls_name.replace("🟢", "").replace("⚫", "").strip()
        filtered = [
            r for r in all_results
            if r.get("student_class", "").strip() == cls_name_clean
        ]
        filtered = filtered[:15]   # Oxirgi 15 ta

        GRADE_ICON = {5: "🏆", 4: "🥈", 3: "🥉", 2: "❌"}
        GRADE_COLOR = {5: "#FFD700", 4: "#42A5F5", 3: "#FFA726", 2: "#EF5350"}

        self._live_table.setRowCount(len(filtered))
        for row, r in enumerate(filtered):
            self._live_table.setRowHeight(row, 34)

            # Vaqt
            dt = (r.get("end_time") or r.get("start_time") or "")
            time_str = dt[11:16] if len(dt) >= 16 else dt[:10]
            t_item = QTableWidgetItem(time_str)
            t_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            t_item.setFont(QFont("Segoe UI", 10))
            t_item.setForeground(QColor("#9E9E9E"))
            self._live_table.setItem(row, 0, t_item)

            # O'quvchi
            name = f"{r.get('student_lastname','')} {r.get('student_name','')}".strip()
            n_item = QTableWidgetItem(name or r.get("student_name", "—"))
            n_item.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
            self._live_table.setItem(row, 1, n_item)

            # Fan (test_name)
            fan_item = QTableWidgetItem(r.get("test_name", "—"))
            fan_item.setFont(QFont("Segoe UI", 10))
            fan_item.setForeground(QColor("#90CAF9"))
            self._live_table.setItem(row, 2, fan_item)

            # To'g'ri / jami
            score_str = f"{r.get('correct_count',0)}/{r.get('total_questions',0)}"
            s_item = QTableWidgetItem(score_str)
            s_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            s_item.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
            is_passed = r.get("is_passed", False)
            s_item.setForeground(QColor("#69F0AE" if is_passed else "#EF9A9A"))
            self._live_table.setItem(row, 3, s_item)

            # Baho
            grade = r.get("grade", 2)
            g_item = QTableWidgetItem(
                f"{GRADE_ICON.get(grade, '?')} {grade}"
            )
            g_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            g_item.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
            g_item.setForeground(QColor(GRADE_COLOR.get(grade, "#9E9E9E")))
            self._live_table.setItem(row, 4, g_item)

    # ── Excel ──────────────────────────────────────────────────────────────────

    def _show_format(self):
        ExcelFormatDialog(self).exec()

    def _import_excel(self):
        if not self._cls_id:
            QMessageBox.warning(self, "Xato", "Avval sinf tanlang!")
            return

        path, _ = QFileDialog.getOpenFileName(
            self, "Excel fayl tanlang", "", "Excel (*.xlsx *.xls)"
        )
        if not path:
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            added, skipped = 0, 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                first = str(row[0]).strip() if row[0] else ""
                last  = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                tg    = str(row[2]).strip() if len(row) > 2 and row[2] else None

                if first and last:
                    api.add_student(self._cls_id, {
                        "first_name": first,
                        "last_name":  last,
                        "parent_telegram_id": tg or None,
                    })
                    added += 1
                else:
                    skipped += 1

            msg = f"✅  {added} ta o'quvchi muvaffaqiyatli yuklandi!"
            if skipped:
                msg += f"\n⚠️  {skipped} ta qator to'liq emas (o'tkazib yuborildi)."
            QMessageBox.information(self, "Import natijasi", msg)
            self._load_students()
            self.refresh()

        except Exception as e:
            QMessageBox.critical(self, "Fayl xatosi", f"Excel faylni o'qishda xato:\n{e}")