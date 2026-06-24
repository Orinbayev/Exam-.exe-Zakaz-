"""
Excel eksport moduli - natijalarni XLSX formatida yaratish.
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
PASS_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMNS = [
    ("Sana", 18),
    ("Ism", 15),
    ("Familiya", 15),
    ("Sinf", 8),
    ("Test nomi", 25),
    ("Jami savollar", 14),
    ("To'g'ri", 10),
    ("Noto'g'ri", 10),
    ("Foiz (%)", 10),
    ("Baho", 8),
    ("Natija", 10),
]


def generate_results_excel(sessions: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Natijalar"

    # Title
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"Smart Exam System - Natijalar ({datetime.now().strftime('%d.%m.%Y')})"
    title_cell.font = Font(bold=True, size=14, color="1565C0")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 25

    # Data rows
    for row_idx, session in enumerate(sessions, start=3):
        date_str = session.end_time.strftime("%d.%m.%Y %H:%M") if session.end_time else "-"
        row_data = [
            date_str,
            session.student_name,
            session.student_lastname,
            session.student_class,
            session.test.name if session.test else "-",
            session.total_questions,
            session.correct_count,
            session.wrong_count,
            round(session.score_percent, 1),
            session.grade or "-",
            "O'tdi" if session.is_passed else "Yiqildi",
        ]

        fill = PASS_FILL if session.is_passed else FAIL_FILL

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx in (6, 7, 8, 9):
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Summary row
    if sessions:
        sum_row = len(sessions) + 3
        ws.merge_cells(f"A{sum_row}:D{sum_row}")
        ws.cell(row=sum_row, column=1).value = "JAMI / O'RTACHA"
        ws.cell(row=sum_row, column=1).font = Font(bold=True)

        scores = [s.score_percent for s in sessions]
        ws.cell(row=sum_row, column=9).value = round(sum(scores) / len(scores), 1)
        ws.cell(row=sum_row, column=9).font = Font(bold=True)

    # Freeze header
    ws.freeze_panes = "A3"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_results_excel_dicts(sessions: list) -> bytes:
    """Dict ro'yxatidan Excel generatsiya qiladi (bot uchun — session yopilgandan keyin)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Natijalar"

    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"Smart Exam System - Natijalar ({datetime.now().strftime('%d.%m.%Y')})"
    title_cell.font = Font(bold=True, size=14, color="1565C0")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 25

    for row_idx, s in enumerate(sessions, start=3):
        date_str = s["end_time"].strftime("%d.%m.%Y %H:%M") if s["end_time"] else "-"
        row_data = [
            date_str,
            s["student_name"],
            s["student_lastname"],
            s["student_class"],
            s["test_name"],
            s["total_questions"],
            s["correct_count"],
            s["wrong_count"],
            round(s["score_percent"], 1),
            s["grade"] or "-",
            "O'tdi" if s["is_passed"] else "Yiqildi",
        ]
        fill = PASS_FILL if s["is_passed"] else FAIL_FILL
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    if sessions:
        sum_row = len(sessions) + 3
        ws.merge_cells(f"A{sum_row}:D{sum_row}")
        ws.cell(row=sum_row, column=1).value = "JAMI / O'RTACHA"
        ws.cell(row=sum_row, column=1).font = Font(bold=True)
        scores = [s["score_percent"] for s in sessions]
        ws.cell(row=sum_row, column=9).value = round(sum(scores) / len(scores), 1)
        ws.cell(row=sum_row, column=9).font = Font(bold=True)

    ws.freeze_panes = "A3"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
