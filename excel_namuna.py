"""
Excel namuna fayl yaratuvchi - savol import uchun shablon.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Savollar"

HEADER_FILL = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = ["Savol matni", "A variant", "B variant", "C variant", "D variant", "To'g'ri javob", "Qiyinlik"]
widths = [50, 25, 25, 25, 25, 14, 12]

for col, (header, width) in enumerate(zip(headers, widths), start=1):
    from openpyxl.utils import get_column_letter
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 25

sample_data = [
    ["2 + 2 nechaga teng?", "3", "4", "5", "6", "B", "easy"],
    ["O'zbekiston poytaxti qaysi shahar?", "Samarqand", "Buxoro", "Toshkent", "Namangan", "C", "easy"],
    ["Suv qanday haroratda qaynaydi?", "80°C", "90°C", "100°C", "110°C", "C", "medium"],
    ["Python qaysi yilda yaratilgan?", "1985", "1991", "1995", "2000", "B", "medium"],
    ["Yerning Quyoshdan masofasi qancha?", "100 mln km", "150 mln km", "200 mln km", "250 mln km", "B", "hard"],
]

for row_idx, row_data in enumerate(sample_data, start=2):
    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row_idx].height = 30

# Izoh
ws.cell(row=len(sample_data) + 3, column=1).value = (
    "IZOH: To'g'ri javob ustunida A, B, C yoki D harfini kiriting. "
    "Qiyinlik: easy, medium yoki hard. "
    "Minimal to'ldirilishi shart: Savol matni, A-D variantlar, To'g'ri javob."
)

import os
save_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "savol_shablon.xlsx")
wb.save(save_path)
print(f"Shablon saqlandi: {save_path}")
